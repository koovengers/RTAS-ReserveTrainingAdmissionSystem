import sys
import json
import os
os.environ.setdefault("QT_API", "pyside6")
import io
import time
from pathlib import Path
import win32com.client
from win32com.client import Dispatch
from openpyxl import Workbook
from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from PIL import Image as PilImage
import pandas as pd
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.drawing.image import Image
from qtpy.QtWidgets import QApplication, QWidget, QVBoxLayout, QPushButton, QFileDialog, QTableWidget, QTableWidgetItem, QLabel, QFrame, QLineEdit, QMessageBox
from qtpy.QtCore import Qt, QPoint
from qtpy.QtWidgets import QDialog, QVBoxLayout, QPushButton, QLabel, QFrame, QLineEdit,QHBoxLayout, QApplication, QWidget, QTableWidget, QTableWidgetItem, QFileDialog, QMessageBox, QCheckBox, QGridLayout, QInputDialog, QProgressDialog
from qtpy.QtGui import QPen, QPainter, QImage, QPixmap, QColor
from qtpy.QtWidgets import *
from qtpy.QtCore import *
from qtpy.QtGui import *

# 중식/교통비(CMS) 관련 클래스
class CheckMealTransportDialog(QDialog):
    def __init__(self, tableWidget, day, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f'{day}일차 열 선택')
        self.tableWidget = tableWidget
        self.day = day  # 일차 변수를 클래스 속성으로 저장
        self.layout = QVBoxLayout(self)

        # 일차에 따라 선택할 수 있는 열 목록 설정
        columns_by_day = {
            1: ["은행", "계좌번호", "예금주", "훈련명칭", "성명", "생년월일", "1일차 표찰", "1일차 중식 미신청 여부", "1일차 중식 신청불가", "1일차 교통비 신청불가", "1일차 훈련비 신청불가"],
            2: ["은행", "계좌번호", "예금주", "훈련명칭", "성명", "생년월일", "2일차 표찰", "2일차 중식 미신청 여부", "2일차 중식 신청불가", "2일차 교통비 신청불가", "2일차 훈련비 신청불가"],
            3: ["은행", "계좌번호", "예금주", "훈련명칭", "성명", "생년월일", "3일차 표찰", "3일차 중식 미신청 여부", "3일차 중식 신청불가", "3일차 교통비 신청불가", "3일차 훈련비 신청불가"],
            4: ["은행", "계좌번호", "예금주", "훈련명칭", "성명", "생년월일", "4일차 표찰", "4일차 중식 미신청 여부", "4일차 중식 신청불가", "4일차 교통비 신청불가", "4일차 훈련비 신청불가"],
            5: ["은행", "계좌번호", "예금주", "훈련명칭", "성명", "생년월일", "5일차 표찰", "5일차 중식 미신청 여부", "5일차 중식 신청불가", "5일차 교통비 신청불가", "5일차 훈련비 신청불가"]
        }

        self.availableColumns = columns_by_day[day]

        self.checkboxes = {}
        for column in self.availableColumns:
            checkBox = QCheckBox(column)
            checkBox.setChecked(self.isColumnAvailable(column))  # 테이블에 열이 존재하는지 확인
            checkBox.setEnabled(self.isColumnAvailable(column))  # 존재하지 않는 열은 선택 불가능
            self.layout.addWidget(checkBox)
            self.checkboxes[column] = checkBox

        self.okButton = QPushButton('확인', self)
        self.okButton.clicked.connect(self.createSheet)
        self.layout.addWidget(self.okButton)

    def isColumnAvailable(self, columnName):
        # 테이블 위젯의 열 이름을 검사하여 해당 열이 존재하는지 확인
        headers = [self.tableWidget.horizontalHeaderItem(i).text() for i in range(self.tableWidget.columnCount())]
        return columnName in headers

    def getSelectedColumns(self):
        # 선택된 열의 목록 반환
        return {column: checkbox.isChecked() for column, checkbox in self.checkboxes.items() if checkbox.isChecked()}

    def createSheet(self):
        base_dir = 'C:/임시폴더'
        base_filename = '임시파일'
        extension = '.xlsx'
        
        # 폴더가 존재하지 않으면 생성
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)

        # 동일한 파일명이 존재할 경우 파일명을 변경
        counter = 0
        while True:
            if counter == 0:
                temp_filename = os.path.join(base_dir, f'{base_filename}{extension}')
            else:
                temp_filename = os.path.join(base_dir, f'{base_filename}_{counter}{extension}')
            if not os.path.exists(temp_filename):
                break
            counter += 1

        day = self.day  # 클래스 속성으로 저장된 day를 사용

        # openpyxl을 사용하여 새로운 워크북 및 시트 생성
        wb = Workbook()
        ws_meal = wb.create_sheet(title=f'{day}일차 중식비')
        ws_transport = wb.create_sheet(title=f'{day}일차 교통비')
        # [추가] 3번째 시트(훈련비) 생성
        ws_training = wb.create_sheet(title=f'{day}일차 훈련비')

        # 기본 생성되는 'Sheet' 시트 제거
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        selected_columns = self.getSelectedColumns()

        # 공통 헤더 생성
        self.createCommonHeaders(ws_meal, selected_columns)
        self.createCommonHeaders(ws_transport, selected_columns)
        # [추가] 훈련비 시트에도 동일한 공통 헤더 생성
        self.createCommonHeaders(ws_training, selected_columns)

        # 각 일차별 중식비 시트의 N1, O1 헤더 설정
        ws_meal['N1'] = f'{day}일차 중식 미신청 여부'
        ws_meal['O1'] = f'{day}일차 중식 신청불가'

        # 각 일차별 교통비 시트의 N1 헤더 설정
        ws_transport['N1'] = f'{day}일차 교통비 신청불가'

        # [추가] 각 일차별 훈련비 시트의 N1 헤더 설정
        ws_training['N1'] = f'{day}일차 훈련비 신청불가'

        # 시트 데이터 채우기
        # [변경] fillSheetData에 pay_type 인자를 넘겨 중식/교통비/훈련비를 구분 처리
        self.fillSheetData(ws_meal, selected_columns, day, pay_type='meal')
        self.fillSheetData(ws_transport, selected_columns, day, pay_type='transport')
        self.fillSheetData(ws_training, selected_columns, day, pay_type='training')

        wb.save(temp_filename)

        # HCell 파일로 저장하기 위한 메서드 호출
        self.saveAsHCell(temp_filename)

        self.accept()

    def createCommonHeaders(self, ws, selected_columns):
        # 공통 헤더 설정
        headers = ["지급구분(*)", "금융기관(*)", "입금계좌(*)", "예금주(*)", "금액(*)", "공급가액", "부가세액", "수령인(*)", "입금통장에 표기(*)", "훈련명칭", "성명", "생년월일", f'{self.day}일차 표찰']
        for col_num, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_num, value=header)

    # [변경] 기존 is_meal 플래그 기반 함수 → pay_type 문자열 기반으로 확장
    #        pay_type: 'meal' / 'transport' / 'training'
    def fillSheetData(self, ws, selected_columns, day, pay_type):
        # 데이터 채우기
        widget = self.tableWidget
        bank_col = self.findColumnIndex("은행")
        account_col = self.findColumnIndex("계좌번호")
        holder_col = self.findColumnIndex("예금주")
        name_col = self.findColumnIndex("성명")
        training_col = self.findColumnIndex("훈련명칭")
        birthdate_col = self.findColumnIndex("생년월일")
        badge_col = self.findColumnIndex(f"{day}일차 표찰")

        # [추가] 지급 구분별로 참조해야 하는 열 인덱스를 분리
        if pay_type == 'meal':
            # 중식비 시트: 중식 미신청 여부 + 중식 신청불가
            flag_col_1 = self.findColumnIndex(f"{day}일차 중식 미신청 여부")
            flag_col_2 = self.findColumnIndex(f"{day}일차 중식 신청불가")
        elif pay_type == 'transport':
            # 교통비 시트: 교통비 신청불가만 사용
            flag_col_1 = self.findColumnIndex(f"{day}일차 교통비 신청불가")
            flag_col_2 = None
        elif pay_type == 'training':
            # [추가] 훈련비 시트: 훈련비 신청불가만 사용
            flag_col_1 = self.findColumnIndex(f"{day}일차 훈련비 신청불가")
            flag_col_2 = None
        else:
            # 예외적인 경우를 대비한 기본값
            flag_col_1 = -1
            flag_col_2 = None

        row = 2
        for widget_row in range(widget.rowCount()):
            # 표찰이 있는(입소한) 인원만 추출
            if badge_col != -1 and widget.item(widget_row, badge_col) and widget.item(widget_row, badge_col).text():
                ws[f'B{row}'] = widget.item(widget_row, bank_col).text() if (bank_col != -1 and widget.item(widget_row, bank_col)) else ""
                ws[f'C{row}'] = widget.item(widget_row, account_col).text() if (account_col != -1 and widget.item(widget_row, account_col)) else ""
                ws[f'D{row}'] = widget.item(widget_row, holder_col).text() if (holder_col != -1 and widget.item(widget_row, holder_col)) else ""
                ws[f'H{row}'] = widget.item(widget_row, name_col).text() if (name_col != -1 and widget.item(widget_row, name_col)) else ""
                ws[f'J{row}'] = widget.item(widget_row, training_col).text() if (training_col != -1 and widget.item(widget_row, training_col)) else ""
                ws[f'K{row}'] = widget.item(widget_row, name_col).text() if (name_col != -1 and widget.item(widget_row, name_col)) else ""
                ws[f'L{row}'] = widget.item(widget_row, birthdate_col).text() if (birthdate_col != -1 and widget.item(widget_row, birthdate_col)) else ""
                ws[f'M{row}'] = widget.item(widget_row, badge_col).text() if (badge_col != -1 and widget.item(widget_row, badge_col)) else ""

                # [변경] 중식비 / 교통비 / 훈련비에 따라 N, O열 기록 방식 분기
                if pay_type == 'meal':
                    # N열: 중식 미신청 여부 / O열: 중식 신청불가
                    ws[f'N{row}'] = widget.item(widget_row, flag_col_1).text() if (flag_col_1 != -1 and widget.item(widget_row, flag_col_1)) else ""
                    ws[f'O{row}'] = widget.item(widget_row, flag_col_2).text() if (flag_col_2 is not None and widget.item(widget_row, flag_col_2)) else ""
                else:
                    # 교통비/훈련비 시트: N열에 각 신청불가 항목만 기입
                    ws[f'N{row}'] = widget.item(widget_row, flag_col_1).text() if (flag_col_1 != -1 and widget.item(widget_row, flag_col_1)) else ""

                row += 1

    def findColumnIndex(self, columnName):
        # 테이블 위젯의 열 인덱스를 찾음
        for i in range(self.tableWidget.columnCount()):
            if self.tableWidget.horizontalHeaderItem(i).text() == columnName:
                return i
        return -1

    def saveAsHCell(self, source_file):
        # 사용자가 파일을 저장할 위치와 이름을 선택할 수 있도록 함
        target_fname, _ = QFileDialog.getSaveFileName(None, 'Save file', './', 'HCell files (*.cell)')
        if target_fname:  # 사용자가 파일 이름을 제공했다면, 처리 계속 진행
            try:
                # 한셀(HCell)로 저장
                hcell = Dispatch('HCell.Application')
                hcell.Visible = True
                workbook = hcell.Workbooks.Open(source_file)
                # HCell 파일 형식으로 저장
                workbook.SaveAs(target_fname, FileFormat=51)
                workbook.Close(False)

                hcell.Quit()

                # 성공 메시지 박스 표시
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Information)
                msg.setWindowTitle("성공")
                msg.setText("파일이 성공적으로 저장되었습니다!")
                msg.setStandardButtons(QMessageBox.Ok)
                msg.exec_()
            except Exception as e:
                print(f"파일 저장 중 오류 발생: {e}")
        else:
            print("파일 저장이 취소되었습니다.")

# 여기까지 중식/교통비(CMS) 클래스 


# 훈련결산 종합하기(동미참) 관련 클래스
class TrainingSummaryDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle('훈련결산 종합하기(동미참)')
        self.layout = QVBoxLayout(self)

        self.filePaths = {}
        self.labels = {}
        self.signature_df = None
        buttonNames = ['최종 전자서명 명부 불러오기', '1일차', '2일차', '3일차', '4일차', '5일차']
        self.dataFrames = {name: pd.DataFrame() for name in buttonNames}

        for name in buttonNames:
            btn = QPushButton(f'{name} 선택', self)
            label = QLabel('파일이 선택되지 않았습니다.', self)
            btn.clicked.connect(lambda _, n=name: self.selectFile(n))
            self.layout.addWidget(btn)
            self.layout.addWidget(label)
            self.filePaths[name] = None
            self.labels[name] = label

        self.summarizeButton = QPushButton('종합', self)
        self.summarizeButton.clicked.connect(self.summarizeData)
        self.layout.addWidget(self.summarizeButton)

    def selectFile(self, name):
        fname, _ = QFileDialog.getOpenFileName(self, '파일 선택', './', 'HCell Files (*.cell)')
        if fname:
            self.labels[name].setText(fname)
            try:
                # os.startfile을 사용하여 파일 열기
                os.startfile(fname)

                # 테이블 형태로 데이터를 입력할 수 있는 창 생성
                tableDialog = QDialog(self)
                tableDialog.setWindowTitle("데이터 확인 및 입력")
                layout = QVBoxLayout(tableDialog)
                
                # 설명 라벨 추가
                instructionsLabel = QLabel("엑셀에서 데이터를 복사하여 아래 텍스트 박스에 붙여넣기 하십시오.")
                layout.addWidget(instructionsLabel)
                
                # 텍스트 에디터 추가
                textEdit = QPlainTextEdit()
                layout.addWidget(textEdit)

                # 텍스트가 변경될 때마다 호출되는 함수 정의
                def validateText():
                    copied_data = textEdit.toPlainText().strip()
                    table_data = [row.split('\t') for row in copied_data.split('\n')]

                    # 1. 각 줄의 열 수를 제한하여 70열을 넘는 데이터를 제거
                    max_columns = 70
                    table_data = [row[:max_columns] for row in table_data]

                    # 2. 줄 수 제한 (2000줄까지만 허용)
                    max_rows = 2000
                    if len(table_data) > max_rows:
                        table_data = table_data[:max_rows]

                    # 수정된 데이터를 다시 텍스트로 변환하여 텍스트 에디터에 입력
                    corrected_text = '\n'.join(['\t'.join(row) for row in table_data])
                    
                    # 텍스트가 다르면 수정된 텍스트로 교체
                    if corrected_text != copied_data:
                        # 이벤트 일시 중지
                        textEdit.blockSignals(True)
                        textEdit.setPlainText(corrected_text)
                        textEdit.blockSignals(False)

                # 텍스트가 변경될 때마다 validateText 함수를 호출
                textEdit.textChanged.connect(validateText)
                
                # 확인 버튼 추가
                confirmButton = QPushButton("확인", tableDialog)
                confirmButton.clicked.connect(lambda: self.TrainingSummaryprocessCopiedData(textEdit, tableDialog, name))
                layout.addWidget(confirmButton)
                
                # 다이얼로그 실행
                tableDialog.exec_()

            except Exception as e:
                QMessageBox.critical(self, "파일 로딩 실패", f"파일을 로딩하는 데 실패했습니다: {e}")
                self.labels[name].setText('파일 로딩 실패')

    def TrainingSummaryprocessCopiedData(self, textEdit, dialog, name):
        copiedText = textEdit.toPlainText()
        if not copiedText:
            QMessageBox.warning(self, "입력 오류", "데이터가 입력되지 않았습니다.")
            return

        if name == '최종 전자서명 명부 불러오기':
            # 데이터를 줄 단위로 분리하고 각 줄을 다시 탭 또는 콤마로 분리하여 2차원 리스트로 변환
            table_data = [line.split('\t') for line in copiedText.split('\n') if line.strip()]

            # 첫 번째 행을 헤더로 사용하여 DataFrame 생성
            df = pd.DataFrame(table_data[1:], columns=table_data[0])

            # ID 열이 비어있는 행 제외
            if 'ID' in df.columns:
                df = df[df['ID'].notna()]

            self.signature_df = df

        else:
            lines = [line.strip() for line in copiedText.split('\n')[14:] if line.strip()]

            table_data = [line.split('\t')[:23] for line in lines]

            # 첫 번째 행을 헤더로 사용하여 DataFrame 생성
            df = pd.DataFrame(table_data[1:], columns=table_data[0])

            # 디버깅 출력: 헤더와 첫 두 행
            print("DataFrame Header:", table_data[0])
            if not df.empty:
                print("First row:", df.iloc[0].to_dict())
                if len(df) > 1:
                    print("Second row:", df.iloc[1].to_dict())

            # G열 또는 H열이 있으면 데이터를 넣어주고 아니면 해당 행을 제외
            df = df[(df.iloc[:, 6].notna()) | (df.iloc[:, 7].notna())]

            self.dataFrames[name] = df

        dialog.accept()

        QMessageBox.information(self, '성공', '파일을 불러오는데 성공했습니다.')

    def summarizeData(self):
        base_dir = 'C:/임시폴더'
        base_filename = '임시파일'
        extension = '.xlsx'
        
        # 폴더가 존재하지 않으면 생성
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)

        # 동일한 파일명이 존재할 경우 파일명을 변경
        counter = 0
        while True:
            if counter == 0:
                temp_filename = os.path.join(base_dir, f'{base_filename}{extension}')
            else:
                temp_filename = os.path.join(base_dir, f'{base_filename}_{counter}{extension}')
            if not os.path.exists(temp_filename):
                break
            counter += 1

        try:

            # 테두리 스타일 설정
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            wb = OpenpyxlWorkbook()
            ws = wb.active
            ws.title = '동원2형결산'

            headers = [
                '훈련명칭+성명+생년월일', '지역대', '훈련명칭', '예비군부대', '성명', '생년월일', '개인차수', '총계획시간'
            ]

            # Create headers
            for idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=idx, value=header)
                ws.cell(row=1, column=idx).alignment = Alignment(horizontal='center')

            if self.signature_df is not None:
                col_mapping = {
                    '지역대': 'B',
                    '훈련명칭': 'C',
                    '예비군부대': 'D',
                    '성명': 'E',
                    '생년월일': 'F',
                    '개인차수': 'G',
                    '총계획시간': 'H'
                }

                for idx, row in enumerate(self.signature_df.itertuples(index=False), start=2):
                    ws[f'A{idx}'] = f'=SUBSTITUTE(C{idx}, "~", "") & E{idx} & F{idx}'
                    ws[f'A{idx}'].alignment = Alignment(horizontal='center')
                    for col_name, col_letter in col_mapping.items():
                        ws[f'{col_letter}{idx}'] = getattr(row, col_name)
                        ws[f'{col_letter}{idx}'].alignment = Alignment(horizontal='center')

            # Hide column A
            ws.column_dimensions['A'].hidden = True

            # 1일차부터 5일차 데이터 작성
            days = ['1일차', '2일차', '3일차', '4일차', '5일차']
            for day in days:
                if day in self.dataFrames and not self.dataFrames[day].empty:
                    df = self.dataFrames[day]
                    sheet = wb.create_sheet(title=day)

                    # 헤더 생성
                    for col_idx, header in enumerate(df.columns, start=2):  # B열부터 시작
                        sheet.cell(row=1, column=col_idx, value=header)
                        # 열 너비 조절
                        max_length = max(len(header), max([len(str(cell)) for cell in df.iloc[:, col_idx - 2]]))
                        sheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

                    # A열 헤더 설정
                    sheet.cell(row=1, column=1, value='훈련명칭+성명+생년월일')
                    sheet.column_dimensions['A'].width = len('훈련명칭+성명+생년월일') + 2

                    for row_idx, row in enumerate(df.itertuples(index=False), start=1):
                        # A열에 함수식 적용
                        cell = sheet.cell(row=row_idx+1, column=1, value=f'=SUBSTITUTE(C{row_idx+1},"~","")&F{row_idx+1}&G{row_idx+1}')
                        cell.alignment = Alignment(horizontal='center')

                        # B열부터 데이터 생성
                        for col_idx, value in enumerate(row, start=2):  # B열부터 시작
                            cell = sheet.cell(row=row_idx+1, column=col_idx, value=value)
                            cell.alignment = Alignment(horizontal='center')

                    # A열 숨기기
                    sheet.column_dimensions['A'].hidden = True

            # 동원2형결산 시트 작성
            summary_sheet = wb['동원2형결산']
            last_row = summary_sheet.max_row

            # 색상 지정 (각 일차별 구분 색상)
            colors = ['FFF4B8', 'E7F7CC', 'D9EBFF', 'F6D9FF', 'FFE3D4']

            # 테두리 스타일 설정
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for day_idx, day in enumerate(days):
                col_offset = day_idx * 11
                color_fill = PatternFill(start_color=colors[day_idx], end_color=colors[day_idx], fill_type="solid")
                
                headers = [
                    f'{day} 입소여부',
                    f'{day} 훈련시간',
                    f'{day} 이수시간',
                    f'{day} 차감시간',
                    f'{day} 차감사유',
                    f'{day} 퇴소유형',
                    f'{day} 미입소유형',
                    f'{day} 중식비',
                    f'{day} 교통비',
                    f'{day} 훈련비',   # [변경] 추가
                    f'{day} 표찰'      # [변경] 표찰은 마지막 열로 이동
                ]

                # 헤더 작성 (시작 열: I열 = 9번 열 고정)
                for col_idx, header in enumerate(headers, start=9 + col_offset):
                    cell = summary_sheet.cell(row=1, column=col_idx, value=header)
                    cell.fill = color_fill
                    summary_sheet.column_dimensions[get_column_letter(col_idx)].width = len(header) + 2

                if day in self.dataFrames and not self.dataFrames[day].empty:
                    for i in range(2, last_row + 1):
                        # [변경 설명]
                        #   - 일자별 시트 구조(훈련결산 저장하기 결과)를 기준으로,
                        #     'day'!A:X 범위에서 필요한 컬럼 인덱스를 다음과 같이 사용:
                        #       11: 입소/미입소(K)
                        #       14: 훈련시간(N)
                        #       18: 이수시간(R)
                        #       15: 차감시간(O)
                        #       16: 차감사유(P)
                        #       19: 퇴소유형(S)
                        #       12: 미입소 유형(L)
                        #       21: 중식비(U)
                        #       22: 교통비(V)
                        #       23: 훈련비(W)
                        #       24: 표찰(X)
                        #   - table_range 도 W까지 → X까지 확대
                        summary_sheet.cell(
                            i, 9 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,11,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,11,0))'
                        )  # 입소여부

                        summary_sheet.cell(
                            i, 10 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,14,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,14,0))'
                        )  # 훈련시간

                        summary_sheet.cell(
                            i, 11 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,18,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,18,0))'
                        )  # 이수시간

                        summary_sheet.cell(
                            i, 12 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,15,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,15,0))'
                        )  # 차감시간

                        summary_sheet.cell(
                            i, 13 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,16,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,16,0))'
                        )  # 차감사유

                        summary_sheet.cell(
                            i, 14 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,19,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,19,0))'
                        )  # 퇴소유형

                        summary_sheet.cell(
                            i, 15 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,12,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,12,0))'
                        )  # 미입소유형

                        summary_sheet.cell(
                            i, 16 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,21,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,21,0))'
                        )  # 중식비

                        summary_sheet.cell(
                            i, 17 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,22,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,22,0))'
                        )  # 교통비

                        summary_sheet.cell(
                            i, 18 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,23,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,23,0))'
                        )  # 훈련비  [변경: 신규 추가]

                        summary_sheet.cell(
                            i, 19 + col_offset
                        ).value = (
                            f'=IF(VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,24,0)="","",'
                            f'VLOOKUP($A{i},\'{day}\'!$A$1:$X$2509,24,0))'
                        )  # 표찰  [변경: 인덱스 24로 조정]

                        for j in range(9 + col_offset, 19 + col_offset):
                            cell = summary_sheet.cell(i, j)
                            cell.alignment = Alignment(horizontal='center')
                            cell.fill = color_fill
                            cell.border = thin_border

            # 동원2형결산 시트에서 B부터 Bk까지의 열 너비를 150픽셀로 설정
            for col_idx in range(2, 64):  # [변경] 59 → 64
                col_letter = get_column_letter(col_idx)
                summary_sheet.column_dimensions[col_letter].width = 150 / 7  # 150픽셀 정도

                for row in range(1, summary_sheet.max_row + 1):
                    cell = summary_sheet.cell(row=row, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')

            # 1일차부터 5일차 시트 설정
            for day in days:
                if day in wb.sheetnames:  # 시트가 존재하는지 확인
                    sheet = wb[day]
                    # [변경] B~W → B~X (df 23열 → 시트 B~X)
                    for col_idx in range(2, 25):  # [변경] 24 → 25 (2~24 = B~X)
                        col_letter = get_column_letter(col_idx)
                        sheet.column_dimensions[col_letter].width = 150 / 7  # 150픽셀 정도

                        for row in range(1, sheet.max_row + 1):
                            cell = sheet.cell(row=row, column=col_idx)
                            cell.border = thin_border
                            cell.alignment = Alignment(horizontal='center')

            wb.save(temp_filename)

            # HCell 파일로 저장하기 위한 메서드 호출
            self.saveAsHCell(temp_filename)

            QMessageBox.information(self, "성공", "파일이 성공적으로 저장되었습니다!")
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {e}")

    def saveAsHCell(self, source_file):
        # 사용자가 파일을 저장할 위치와 이름을 선택할 수 있도록 함
        target_fname, _ = QFileDialog.getSaveFileName(None, 'Save file', './', 'HCell files (*.cell)')
        if target_fname:  # 사용자가 파일 이름을 제공했다면, 처리 계속 진행
            try:
                # 한셀(HCell)로 저장
                hcell = Dispatch('HCell.Application')
                hcell.Visible = True
                workbook = hcell.Workbooks.Open(source_file)
                # HCell 파일 형식으로 저장
                workbook.SaveAs(target_fname, FileFormat=51)
                workbook.Close(False)

                workbook.Save()
                workbook.Close(False)
                hcell.Quit()

                print("파일이 성공적으로 저장되었습니다!")
            except Exception as e:
                print(f"파일 저장 중 오류 발생: {e}")
        else:
            print("파일 저장이 취소되었습니다.")

# 여기까지 훈련결산 종합하기(동미참) 클래스




# 훈련결산 저장하기 관련 클래스
class SaveColumnSelectorDialog(QDialog):
    def __init__(self, tableWidget, day, parent=None):
        super().__init__(parent)
        self.setWindowTitle(f'{day}일차 열 선택')
        self.tableWidget = tableWidget
        self.day = day  # day 변수를 클래스 속성으로 저장
        self.layout = QVBoxLayout(self)

        # 일차에 따라 사용자가 선택할 수 있는 열 목록
        columns_by_day = {
            1: ["지역대", "훈련시작일차", "훈련종료일차", "훈련명칭", "예비군부대", "계급", "성명", "생년월일", "훈련유형", "개인차수", "총계획시간",
                "1일차 훈련시간", '1일차 차감시간', '1일차 차감사유', "1일차 표찰", "1일차 입소서명", "1일차 퇴소서명"],
            2: ["지역대", "훈련시작일차", "훈련종료일차", "훈련명칭", "예비군부대", "계급", "성명", "생년월일", "훈련유형", "개인차수", "총계획시간",
                "2일차 훈련시간", '2일차 차감시간', '2일차 차감사유', "2일차 표찰", "2일차 입소서명", "2일차 퇴소서명"],
            3: ["지역대", "훈련시작일차", "훈련종료일차", "훈련명칭", "예비군부대", "계급", "성명", "생년월일", "훈련유형", "개인차수", "총계획시간",
                "3일차 훈련시간", '3일차 차감시간', '3일차 차감사유', "3일차 표찰", "3일차 입소서명", "3일차 퇴소서명"],
            4: ["지역대", "훈련시작일차", "훈련종료일차", "훈련명칭", "예비군부대", "계급", "성명", "생년월일", "훈련유형", "개인차수", "총계획시간",
                "4일차 훈련시간", '4일차 차감시간', '4일차 차감사유', "4일차 표찰", "4일차 입소서명", "4일차 퇴소서명"],
            5: ["지역대", "훈련시작일차", "훈련종료일차", "훈련명칭", "예비군부대", "계급", "성명", "생년월일", "훈련유형", "개인차수", "총계획시간",
                "5일차 훈련시간", '5일차 차감시간', '5일차 차감사유', "5일차 표찰", "5일차 입소서명", "5일차 퇴소서명"]
        }

        self.availableColumns = columns_by_day[day]

        self.checkboxes = {}
        for column in self.availableColumns:
            checkBox = QCheckBox(column)
            checkBox.setChecked(self.isColumnAvailable(column))  # 테이블에 열이 존재하는지 확인
            checkBox.setEnabled(self.isColumnAvailable(column))  # 존재하지 않는 열은 선택 불가능
            self.layout.addWidget(checkBox)
            self.checkboxes[column] = checkBox

        self.okButton = QPushButton('확인', self)
        self.okButton.clicked.connect(self.createSheet)
        self.layout.addWidget(self.okButton)

    def isColumnAvailable(self, columnName):
        # 테이블 위젯의 열 이름을 검사하여 해당 열이 존재하는지 확인
        headers = [self.tableWidget.horizontalHeaderItem(i).text() for i in range(self.tableWidget.columnCount())]
        return columnName in headers

    def getSelectedColumns(self):
        # 선택된 열의 목록 반환
        return {column: checkbox.isChecked() for column, checkbox in self.checkboxes.items() if checkbox.isChecked()}

    def createSheet(self):
        base_dir = 'C:/임시폴더'
        base_filename = '임시파일'
        extension = '.xlsx'
        
        # 폴더가 존재하지 않으면 생성
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)

        # 동일한 파일명이 존재할 경우 파일명을 변경
        counter = 0
        while True:
            if counter == 0:
                temp_filename = os.path.join(base_dir, f'{base_filename}{extension}')
            else:
                temp_filename = os.path.join(base_dir, f'{base_filename}_{counter}{extension}')
            if not os.path.exists(temp_filename):
                break
            counter += 1

        day = self.day  # 클래스 속성으로 저장된 day를 사용

        # openpyxl을 사용하여 새로운 워크북 및 시트 생성
        wb = Workbook()
        ws = wb.active
        ws.title = f'{day}일차'

        # 기본 생성되는 'Sheet' 시트 제거
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']

        # 열 헤더를 가져오기
        headers = [self.tableWidget.horizontalHeaderItem(i).text() for i in range(self.tableWidget.columnCount())]

        # 훈련시작일차와 훈련종료일차 열 인덱스 찾기
        start_day_idx = headers.index("훈련시작일차")
        end_day_idx = headers.index("훈련종료일차")

        # 유효한 행 필터링
        valid_rows = []
        for row in range(self.tableWidget.rowCount()):
            start_day = self.tableWidget.item(row, start_day_idx)
            end_day = self.tableWidget.item(row, end_day_idx)
            if start_day and end_day and start_day.text() and end_day.text():
                try:
                    start_day_value = int(start_day.text())
                    end_day_value = int(end_day.text())
                    if day >= start_day_value and day <= end_day_value:
                        valid_rows.append(row)
                except ValueError:
                    continue
        # A열
        ws['A7'] = '훈련명칭+성명+생년월일'
        for idx, row in enumerate(valid_rows, start=8):
            ws[f'A{idx}'] = f'=SUBSTITUTE(D{idx}, "~", "") & G{idx} & H{idx}'

        # B열
        ws['B7'] = '성명+생년월일'
        for idx, row in enumerate(valid_rows, start=8):
            ws[f'B{idx}'] = f'=G{idx}&H{idx}'

        # C열부터 V열까지 각 열에 데이터 설정
        column_mappings = {
            'C': '지역대', 'D': '훈련명칭', 'E': '예비군부대', 'F': '계급', 'G': '성명', 'H': '생년월일',
            'I': '훈련유형', 'J': '개인차수', 'K': '총계획시간', 'L': '입소/미입소', 'M': '미입소 유형', 'N': '무단불참 여부',
            'O': f'{day}일차 훈련시간', 'P': f'{day}일차 차감시간', 'Q': f'{day}일차 차감사유', 'R': '퇴소예정시간', 'S': '훈련이수시간', 'T': '퇴소유형', 'U': '잔여시간',
            'V': '중식비 지급여부', 'W': '교통비 지급여부', 'X': '훈련비 지급여부', 'Y': f'{day}일차 표찰'
        }

        for col, header in column_mappings.items():
            ws[f'{col}7'] = header
            if header in headers:
                col_idx = headers.index(header)
                for idx, row in enumerate(valid_rows, start=8):
                    cell_value = self.tableWidget.item(row, col_idx).text() if self.tableWidget.item(row, col_idx) else ""
                    ws[f'{col}{idx}'] = cell_value

        # L열에 입소/미입소 값 설정
        for row in range(8, ws.max_row + 1):
            if ws[f'Y{row}'].value:
                ws[f'L{row}'] = '입소'
            else:
                ws[f'L{row}'] = '미입소'

        # N열에 수식 설정
        ws['N7'] = '무단불참 여부'
        for idx in range(8, ws.max_row + 1):
            ws[f'N{idx}'] = f'=IF(ISBLANK(L{idx}),"",IF(AND(L{idx}="미입소",M{idx}="음주미입소"),"",IF(AND(L{idx}="미입소",M{idx}="신검불합"),"",IF(AND(L{idx}="미입소",M{idx}="전출"),"",IF(AND(L{idx}="미입소",M{idx}="신고불참"),"",IF(AND(L{idx}="미입소",M{idx}="연기"),"",IF(AND(L{idx}="미입소",M{idx}="미교부"),"",IF(AND(L{idx}="미입소",M{idx}="보류"),"",IF(OR(R{idx}="정상퇴소",R{idx}="개인업무",R{idx}="건강악화"),"",IF(L{idx}="입소","","무단불참"))))))))))'

        # R열에 수식 설정
        ws['R7'] = '퇴소예정시간'
        for idx in range(8, ws.max_row + 1):
            ws[f'R{idx}'] = f'=IF(O{idx}-P{idx}=8,"18시",IF(O{idx}-P{idx}=7,"17시",IF(O{idx}-P{idx}=6,"16시",IF(O{idx}-P{idx}=5,"15시",IF(O{idx}-P{idx}=4,"14시",IF(O{idx}-P{idx}=3,"12시",IF(O{idx}-P{idx}=2,"11시",IF(O{idx}-P{idx}=1,"10시","오류"))))))))'

        # U열에 수식 설정
        ws['U7'] = '잔여시간'
        for idx in range(8, ws.max_row + 1):
            ws[f'U{idx}'] = f'=SUM(O{idx}-P{idx}-S{idx})'

        # V열에 수식 설정
        ws['V7'] = '중식비 지급여부'
        for row in range(8, ws.max_row + 1):
            ws[f'V{row}'] = (
                f'=IFERROR('
                f'IF(VLOOKUP(B{row},$Z$12:$AF$1013,5,0)=0,"",'
                f'VLOOKUP(B{row},$Z$12:$AF$1013,5,0)'
                f'),"")'
            )

        # W열에 수식 설정
        ws['W7'] = '교통비 지급여부'
        for row in range(8, ws.max_row + 1):
            ws[f'W{row}'] = (
                f'=IFERROR('
                f'IF(VLOOKUP(B{row},$Z$12:$AF$1013,6,0)=0,"",'
                f'VLOOKUP(B{row},$Z$12:$AF$1013,6,0)'
                f'),"")'
            )

        # x열에 수식 설정
        ws['X7'] = '훈련비 지급여부'
        for row in range(8, ws.max_row + 1):
            ws[f'X{row}'] = (
                f'=IFERROR('
                f'IF(VLOOKUP(B{row},$Z$12:$AF$1013,7,0)=0,"",'
                f'VLOOKUP(B{row},$Z$12:$AF$1013,7,0)'
                f'),"")'
            )


        # AB열부터 AT열까지 설정
        ws['AB7'] = '하령인원'
        ws['AB8'] = '=COUNTIF($H$8:$H$5000,">=1")'

        ws['AC7'] = '훈련대상'
        ws['AC8'] = '=SUM(AB8-AK8-AL8-AM8-AP8)'

        ws['AD7'] = '입소인원'
        ws['AD8'] = '=COUNTIF($L$8:$L$5000,"입소")'

        ws['AE7'] = '미입소인원'
        ws['AE8'] = '=COUNTIF($L$8:$L$5000,"미입소")'

        ws['AF7'] = '입소율'
        ws['AF8'] = '=AD8/AC8*100'

        ws['AH7'] = '계'
        ws['AH8'] = '=SUM(AI8:AP8)'

        ws['AI7'] = '신고불참'
        ws['AI8'] = '=COUNTIF($M$8:$M$5000,"신고불참")'

        ws['AJ7'] = '신검불합'
        ws['AJ8'] = '=COUNTIF($M$8:$M$5000,"신검불합")'

        ws['AK7'] = '전출'
        ws['AK8'] = '=COUNTIF($M$8:$M$5000,"전출")'

        ws['AL7'] = '연기'
        ws['AL8'] = '=COUNTIF($M$8:$M$5000,"연기")'

        ws['AM7'] = '미교부'
        ws['AM8'] = '=COUNTIF($M$8:$M$5000,"미교부")'

        ws['AN7'] = '무단불참'
        ws['AN8'] = '=COUNTIF($N$8:$N$5000,"무단불참")'

        ws['AO7'] = '음주미입소'
        ws['AO8'] = '=COUNTIF($M$8:$M$5000,"음주미입소")'

        ws['AP7'] = '보류'
        ws['AP8'] = '=COUNTIF($M$8:$M$5000,"보류")'

        ws['AR7'] = '계'
        ws['AR8'] = '=SUM(AS8:AT8)'

        ws['AS7'] = '정상퇴소'
        ws['AS8'] = '=COUNTIF($T$8:$T$5000,"정상퇴소")'

        ws['AT7'] = '조기퇴소'
        ws['AT8'] = '=COUNTIF($T$8:$T$5000,"건강악화")+COUNTIF($T$8:$T$5000,"개인업무")'

        # 각 열의 너비 설정
        for col in ws.iter_cols(min_row=7, max_row=7):
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 7)
            ws.column_dimensions[column].width = adjusted_width

        # M열의 너비를 두 배로 설정
        current_width = ws.column_dimensions['M'].width
        ws.column_dimensions['M'].width = current_width * 2            

        # A열과 B열 숨기기
        ws.column_dimensions['A'].hidden = True
        ws.column_dimensions['B'].hidden = True

        # C1:L6까지 셀 통합하고 '자동계산 입력금지'이라는 단어 넣기
        ws.merge_cells('C1:L6')
        ws['C1'] = '자동계산 입력금지'

        # M1:M5까지 셀 통합하고 내용 입력
        ws.merge_cells('M1:M5')
        ws['M1'] = "연기, 보류, 미교부\n신고불참, 음주미입소\n신검불합, 전출\n\n위 8가지 사유만 입력"

        # M6에 '입소시 입력(Chcek!)'라는 단어 넣기
        ws['M6'] = '입소시 입력(Chcek!)'

        # N1:R6까지 셀 통합하고 '자동계산 입력금지'이라는 단어 넣기
        ws.merge_cells('N1:R6')
        ws['N1'] = '자동계산 입력금지'

        # S1:T6까지 셀 통합하고 '퇴소시 입력(Chcek!)'라는 단어 넣기
        ws.merge_cells('S1:T6')
        ws['S1'] = '퇴소유형\n(정상퇴소, 건강악화,\n개인업무)\n\n퇴소시 입력(Chcek!)'

        # U1:Y6까지 셀 통합하고 '자동계산 입력금지'이라는 단어 넣기
        ws.merge_cells('U1:Y6')
        ws['U1'] = '자동계산 입력금지'

        #   - 상단 요약 박스 병합 영역도 AA~AS → AB~AT로 이동
        #   - 기존: AA1:AS5 → 변경: AB1:AT5
        ws.merge_cells('AB1:AT5')
        ws['AB1'] = '입/퇴소 현황'

        #   - 기존: AA6:AE6 → 변경: AB6:AF6
        ws.merge_cells('AB6:AF6')
        ws['AB6'] = '입/퇴소 총계'

        #   - 기존: AG6:AO6 → 변경: AH6:AP6
        ws.merge_cells('AH6:AP6')
        ws['AH6'] = '미입소 유형'

        #   - 기존: AQ6:AS6 → 변경: AR6:AT6
        ws.merge_cells('AR6:AT6')
        ws['AR6'] = '퇴소 유형'

        # Z열 헤더 설정 (표시 텍스트는 동일)
        ws['Z11'] = '성명+생년월일'

        # Z열 함수식 적용
        for row in range(12, 1013):
            ws[f'Z{row}'] = f'=AB{row}&AC{row}'

        # 각 열의 헤더 설정
        ws['AB11'] = '성명'
        ws['AC11'] = '생년월일'
        ws['AD11'] = '중식비'
        ws['AE11'] = '교통비'
        ws['AF11'] = '훈련비'

        # '중식비/교통비/훈련비 입력'이라는 단어 넣기
        ws['AB10'] = '중식/교통/훈련비 입력'         

        # Z열 숨기기
        ws.column_dimensions['Z'].hidden = True

        # 시트 전체 모든 셀의 높이와 너비를 가운데 정렬
        alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        fill_faebd7 = PatternFill(start_color="FAEBD7", end_color="FAEBD7", fill_type="solid")

        fill_e6ecf0 = PatternFill(start_color="E6ECF0", end_color="E6ECF0", fill_type="solid")

        # 지정된 열의 전체 배경색 설정
        columns = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'N', 'O', 'P', 'Q', 'R', 'U', 'V', 'W', 'X', 'Y']
        for col in columns:
            for row in range(1, ws.max_row + 1):
                ws[f'{col}{row}'].fill = fill_faebd7

        # M1:M7 셀의 배경색 설정
        for row in range(1, 8):
            ws[f'M{row}'].fill = fill_e6ecf0

        # S1:T7 셀의 배경색 설정
        for col in ['S', 'T']:
            for row in range(1, 8):
                ws[f'{col}{row}'].fill = fill_e6ecf0

        # AA10:AD11 셀의 배경색 설정
        for row in range(10, 12):
            for col in range(28, 33):  # 28=AB, 29=AC, 30=AD, 31=AE, 32=AF
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_e6ecf0             

        # 지정된 셀의 배경색 설정
        cells = [
            'AB1', 'AB6', 'AB7', 'AB8', 'AC7', 'AC8', 'AD7', 'AD8', 'AE7', 'AE8', 
            'AH6', 'AH7', 'AH8', 'AI7', 'AI8', 'AJ7', 'AJ8', 'AK7', 'AK8', 
            'AL7', 'AL8', 'AM7', 'AM8', 'AN7', 'AN8', 'AO7', 'AO8', 'AP7', 'AP8',
            'AR6', 'AR7', 'AR8', 'AS7', 'AS8', 'AT7', 'AT8'
        ]
        for cell in cells:
            ws[cell].fill = fill_faebd7

        # 테두리 스타일 정의
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # a1:y2000 범위의 테두리 설정
        for row in ws.iter_rows(min_row=1, max_row=2000, min_col=1, max_col=25):
            for cell in row:
                cell.border = thin_border

        # Ab1:At6 범위의 테두리 설정
        for row in ws.iter_rows(min_row=1, max_row=6, min_col=28, max_col=46):
            for cell in row:
                cell.border = thin_border

        # Ab6:af6 범위의 테두리 설정
        for row in ws.iter_rows(min_row=6, max_row=6, min_col=28, max_col=32):
            for cell in row:
                cell.border = thin_border

        # Ah6:Ap6 범위의 테두리 설정
        for row in ws.iter_rows(min_row=6, max_row=6, min_col=34, max_col=42):
            for cell in row:
                cell.border = thin_border

        # Ar6:At6 범위의 테두리 설정
        for row in ws.iter_rows(min_row=6, max_row=6, min_col=44, max_col=46):
            for cell in row:
                cell.border = thin_border

        # Ab10:Af1011 범위의 테두리 설정
        for row in ws.iter_rows(min_row=10, max_row=1011, min_col=28, max_col=32):
            for cell in row:
                cell.border = thin_border         

        # 개별 셀의 테두리 설정
        individual_cells = [
            'AB7', 'AB8', 'AC7', 'AC8', 'AD7', 'AD8', 'AE7', 'AE8', 'AF7', 'AF8', 
            'AH7', 'AH8', 'AI7', 'AI8', 'AJ7', 'AJ8', 'AK7', 'AK8', 
            'AL7', 'AL8', 'AM7', 'AM8', 'AN7', 'AN8', 'AO7', 'AO8',
            'AP7', 'AP8', 'AR7', 'AR8', 'AS7', 'AS8', 'AT7', 'AT8'
        ]
        for cell in individual_cells:
            ws[cell].border = thin_border

        # 빨간색 글씨와 굵게 설정
        red_bold_font = Font(color="FF0000", bold=True)

        # 파란색 글씨와 굵게 설정
        blue_bold_font = Font(color="0000FF", bold=True)

        # C1, N1, S1 셀의 글씨를 빨간색으로 굵게 설정
        ws['C1'].font = red_bold_font
        ws['N1'].font = red_bold_font
        ws['U1'].font = red_bold_font

        # M1, M6, Q1 셀의 글씨를 파란색으로 굵게 설정
        ws['M1'].font = blue_bold_font
        ws['M6'].font = blue_bold_font
        ws['S1'].font = blue_bold_font

        wb.save(temp_filename)

        # HCell 파일로 저장하기 위한 메서드 호출
        self.saveAsHCell(temp_filename)

        self.accept()

    def saveAsHCell(self, source_file):
        # 사용자가 파일을 저장할 위치와 이름을 선택할 수 있도록 함
        target_fname, _ = QFileDialog.getSaveFileName(None, 'Save file', './', 'HCell files (*.cell)')
        if target_fname:  # 사용자가 파일 이름을 제공했다면, 처리 계속 진행
            try:
                # 한셀(HCell)로 저장
                hcell = Dispatch('HCell.Application')
                hcell.Visible = True
                workbook = hcell.Workbooks.Open(source_file)
                # HCell 파일 형식으로 저장
                workbook.SaveAs(target_fname, FileFormat=51)
                workbook.Close(False)

                workbook.Save()
                workbook.Close(False)
                hcell.Quit()

                # 파일이 성공적으로 저장되었음을 알리는 메시지 박스
                QMessageBox.information(None, "성공", "파일이 성공적으로 저장되었습니다!")
            except Exception as e:
                print(f"파일 저장 중 오류 발생: {e}")
        else:
            print("파일 저장이 취소되었습니다.")

# 훈련결산 저장하기 관련 클래스 여기까지



# 전자서명 명부 종합하기 버튼을 클래스
class MergeFilesDialog(QDialog):
    def __init__(self, parent=None, signatureSavePath=None):
        super(MergeFilesDialog, self).__init__(parent)
        self.setWindowTitle('전자서명 명부 종합하기')
        self.layout = QVBoxLayout(self)
        
        self.filePaths = {}
        self.buttonNames = ['원본', '첫번째 파일', '두번째 파일', '세번째 파일', '네번째 파일']
        self.dataFrames = {name: None for name in self.buttonNames}

        # 서명 이미지 경로 추가
        self.signatureSavePath = signatureSavePath

        for name in self.buttonNames:
            btn = QPushButton(f'{name} 선택', self)
            label = QLabel('파일이 선택되지 않았습니다.', self)
            btn.clicked.connect(lambda _, n=name, l=label: self.selectFile(n, l))
            self.layout.addWidget(btn)
            self.layout.addWidget(label)
            self.filePaths[name] = label
        
        # 종합 버튼 추가
        self.summarizeButton = QPushButton('종합', self)
        self.summarizeButton.clicked.connect(self.summarizeData)
        self.layout.addWidget(self.summarizeButton)

        # 데이터 표시 테이블 추가
        self.tableWidget = QTableWidget()
        self.layout.addWidget(self.tableWidget)

        # 저장하기 버튼 추가
        self.saveButton = QPushButton('저장하기', self)
        self.saveButton.clicked.connect(self.saveData)
        self.layout.addWidget(self.saveButton)        

    def selectFile(self, name, label):
        fname, _ = QFileDialog.getOpenFileName(self, '파일 선택', './', 'HCell Files (*.cell)')
        if fname:
            label.setText(fname)
            # 한셀을 통해 파일 읽기
            try:
                # os.startfile을 사용하여 파일 열기
                os.startfile(fname)

                # 테이블 형태로 데이터를 입력할 수 있는 창 생성
                tableDialog = QDialog(self)
                tableDialog.setWindowTitle("데이터 확인 및 입력")
                layout = QVBoxLayout(tableDialog)
                
                # 설명 라벨 추가
                instructionsLabel = QLabel("엑셀에서 데이터를 복사하여 아래 텍스트 박스에 붙여넣기 하십시오.")
                layout.addWidget(instructionsLabel)
                
                # 텍스트 에디터 추가
                textEdit = QPlainTextEdit()
                layout.addWidget(textEdit)

                # 텍스트가 변경될 때마다 호출되는 함수 정의
                def validateText():
                    copied_data = textEdit.toPlainText().strip()
                    table_data = [row.split('\t') for row in copied_data.split('\n')]

                    # 1. 각 줄의 열 수를 제한하여 70열을 넘는 데이터를 제거
                    max_columns = 70
                    table_data = [row[:max_columns] for row in table_data]

                    # 2. 줄 수 제한 (2000줄까지만 허용)
                    max_rows = 2000
                    if len(table_data) > max_rows:
                        table_data = table_data[:max_rows]

                    # 수정된 데이터를 다시 텍스트로 변환하여 텍스트 에디터에 입력
                    corrected_text = '\n'.join(['\t'.join(row) for row in table_data])
                    
                    # 텍스트가 다르면 수정된 텍스트로 교체
                    if corrected_text != copied_data:
                        # 이벤트 일시 중지
                        textEdit.blockSignals(True)
                        textEdit.setPlainText(corrected_text)
                        textEdit.blockSignals(False)

                # 텍스트가 변경될 때마다 validateText 함수를 호출
                textEdit.textChanged.connect(validateText)
                
                # 확인 버튼 추가
                confirmButton = QPushButton("확인", tableDialog)
                confirmButton.clicked.connect(lambda: self.mergeprocessCopiedData(textEdit, tableDialog, name))
                layout.addWidget(confirmButton)
                
                # 다이얼로그 실행
                tableDialog.exec_()

            except Exception as e:
                QMessageBox.critical(self, "파일 로딩 실패", f"파일을 로딩하는 데 실패했습니다: {e}")
                label.setText('파일 로딩 실패')

    def mergeprocessCopiedData(self, textEdit, dialog, name):
        copiedText = textEdit.toPlainText()
        if not copiedText:
            QMessageBox.warning(self, "입력 오류", "데이터가 입력되지 않았습니다.")
            return
        
        # 데이터를 줄 단위로 분리하고 각 줄을 다시 탭으로 분리하여 2차원 리스트로 변환
        table_data = [line.split('\t') for line in copiedText.split('\n') if line.strip()]

        if len(table_data) == 0:
            QMessageBox.warning(self, "입력 오류", "데이터가 유효하지 않습니다.")
            return

        # 첫 번째 행을 헤더로 사용하여 DataFrame 생성
        df = pd.DataFrame(table_data[1:], columns=table_data[0])
        
        # ID 열이 비어있는 행 제외
        if 'ID' in df.columns:
            df = df[df['ID'].notna()]

        self.dataFrames[name] = df

        # 1행과 2행 데이터를 디버깅 출력
        if len(df) > 0:
            print("1행 데이터:", df.iloc[0].to_dict())
        if len(df) > 1:
            print("2행 데이터:", df.iloc[1].to_dict())

        dialog.accept()

        QMessageBox.information(self, '성공', '파일을 불러오는데 성공했습니다.') 

    def summarizeData(self):
        try:
            print("데이터 병합 시작...")  # 병합 시작 로그 출력
            
            base_df = self.dataFrames.get('원본')  # '원본' 데이터 프레임 가져오기
            if base_df is None:
                QMessageBox.critical(self, "병합 실패", "'원본' 데이터가 없습니다.")
                return  # 함수 종료

            if 'ID' not in base_df.columns:  # 'ID' 열이 없는 경우
                # 기본 선택값 설정
                name_column_default = '성명' if '성명' in base_df.columns else base_df.columns[0]
                birth_column_default = '생년월일' if '생년월일' in base_df.columns else base_df.columns[0]
                type_column_default = '훈련유형' if '훈련유형' in base_df.columns else base_df.columns[0]
                count_column_default = '개인차수' if '개인차수' in base_df.columns else base_df.columns[0]

                # 성명, 생년월일, 훈련유형, 개인차수 열을 확인해서 ID 열 생성
                name_column, ok1 = QInputDialog.getItem(self, "성명 열 선택", "성명 열을 선택하세요:", base_df.columns.tolist(), base_df.columns.tolist().index(name_column_default), False)
                birth_column, ok2 = QInputDialog.getItem(self, "생년월일 열 선택", "생년월일 열을 선택하세요:", base_df.columns.tolist(), base_df.columns.tolist().index(birth_column_default), False)
                type_column, ok3 = QInputDialog.getItem(self, "훈련유형 열 선택", "훈련유형 열을 선택하세요:", base_df.columns.tolist(), base_df.columns.tolist().index(type_column_default), False)
                count_column, ok4 = QInputDialog.getItem(self, "개인차수 열 선택", "개인차수 열을 선택하세요:", base_df.columns.tolist(), base_df.columns.tolist().index(count_column_default), False)

                if not all([ok1, ok2, ok3, ok4]):
                    QMessageBox.warning(self, "선택 오류", "모든 열을 선택해야 합니다.")
                    return

                base_df['ID'] = base_df[name_column] + "_" + base_df[birth_column] + "_" + base_df[type_column] + "_" + base_df[count_column]
                # ID 열을 첫 번째 열로 이동
                cols = ['ID'] + [col for col in base_df.columns if col != 'ID']
                base_df = base_df[cols]

            merge_keys = ['ID']  # 병합 키로 'ID' 열 설정
            file_dfs = [self.dataFrames.get(f'{i}번째 파일') for i in ['첫', '두', '세', '네']]  # '첫번째 파일', '두번째 파일', '세번째 파일', '네번째 파일' 데이터 프레임 가져오기
            
            # 데이터 병합
            current_df = base_df.copy()
            for i, df in enumerate(file_dfs):
                if df is not None:
                    if 'ID' not in df.columns:
                        QMessageBox.critical(self, "병합 실패", f"'{i+1}번째 파일' 데이터에 'ID' 열이 없습니다.")
                        return
                    current_df = self.mergeDataFrames(current_df, df, merge_keys)
                    print(f"{i+1}번째 파일 데이터 병합 완료.")  # 병합 완료 로그 출력

            print("모든 데이터 병합 완료. 최종 처리 단계 진입...")  # 모든 데이터 병합 완료 로그 출력
            
            self.displayDataInTable(current_df, base_df['ID'].tolist())

        except Exception as e:
            QMessageBox.critical(self, "병합 실패", f"데이터 병합에 실패했습니다: {e}")  # 예외 발생 시 에러 메시지 표시

    def mergeDataFrames(self, base_df, new_df, merge_keys):
        merged_df = pd.merge(base_df, new_df, on=merge_keys, how='outer', suffixes=('', '_new'))
        for col in base_df.columns:
            if col in merge_keys:
                continue
            new_col = col + '_new'
            if new_col in merged_df.columns:
                merged_df[col] = merged_df.apply(lambda row: self.resolveConflict(row[col], row[new_col]), axis=1)
                merged_df.drop(columns=[new_col], inplace=True)
        return merged_df

    def resolveConflict(self, base_value, new_value):
        if pd.isna(base_value) and pd.notna(new_value):
            return new_value
        if pd.notna(new_value) and base_value != new_value:
            if pd.isna(base_value) or base_value == '':
                return new_value
            return new_value if new_value != '' else base_value
        return base_value

    def displayDataInTable(self, df, id_order):
        # id_order에 있는 행과 없는 행을 분리합니다.
        df_ordered = df[df['ID'].isin(id_order)].set_index('ID').loc[id_order].reset_index()
        df_not_ordered = df[~df['ID'].isin(id_order)]

        # 정렬된 데이터프레임과 정렬되지 않은 데이터프레임을 병합합니다.
        df_final = pd.concat([df_ordered, df_not_ordered], ignore_index=True)

        self.tableWidget.setColumnCount(len(df_final.columns))
        self.tableWidget.setRowCount(len(df_final))
        self.tableWidget.setHorizontalHeaderLabels(df_final.columns)

        for i, row in df_final.iterrows():
            for j, val in enumerate(row):
                if pd.isna(val):
                    val = ''  # NaN 값을 빈 문자열로 대체
                elif isinstance(val, float) and val.is_integer():
                    val = str(int(val))  # .0이 붙은 정수를 정수로 변환
                else:
                    val = str(val)
                self.tableWidget.setItem(i, j, QTableWidgetItem(val))

    def saveData(self):
        try:
            final_df = self.getTableData()
            temp_filename = self.saveDataAsExcel(final_df)
            if temp_filename:
                self.processFileWithHCell(temp_filename)
                print("데이터 저장 완료.")
        except Exception as e:
            QMessageBox.critical(self, "저장 실패", f"데이터 저장에 실패했습니다: {e}")

    def getTableData(self):
        rows = self.tableWidget.rowCount()
        cols = self.tableWidget.columnCount()
        headers = [self.tableWidget.horizontalHeaderItem(i).text() for i in range(cols)]
        
        data = []
        for row in range(rows):
            row_data = []
            for col in range(cols):
                item = self.tableWidget.item(row, col)
                row_data.append(item.text() if item else '')
            data.append(row_data)

        return pd.DataFrame(data, columns=headers)

    def saveDataAsExcel(self, df):
        base_dir = 'C:/임시폴더'
        base_filename = '임시파일'
        extension = '.xlsx'
        
        # 폴더가 존재하지 않으면 생성
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)

        # 동일한 파일명이 존재할 경우 파일명을 변경
        counter = 0
        while True:
            if counter == 0:
                temp_filename = os.path.join(base_dir, f'{base_filename}{extension}')
            else:
                temp_filename = os.path.join(base_dir, f'{base_filename}_{counter}{extension}')
            if not os.path.exists(temp_filename):
                break
            counter += 1

        try:
            wb = OpenpyxlWorkbook()
            ws = wb.active

            for j, col in enumerate(df.columns):
                ws.cell(row=1, column=j+1).value = col

            for i, row in enumerate(df.itertuples(index=False), start=2):
                for j, value in enumerate(row, start=1):
                    cell = ws.cell(row=i, column=j)


                    # 이미지 경로 변경
                    if isinstance(value, str) and (value.endswith('.png') or value.endswith('.gif')):
                        # 경로에서 파일명 추출
                        filename = os.path.basename(value)
                        # 새로운 경로로 파일명 결합
                        new_path = os.path.join(self.signatureSavePath, filename)
                        value = new_path  # 이미지 경로를 무조건 변경된 경로로 설정

                    cell.value = value

                    # 이미지 파일 경로가 있는 경우 이미지를 셀에 배치
                    if isinstance(value, str) and value.lower().endswith(('.png', '.gif')) and os.path.exists(value):
                        img = Image(value)

                        # 이미지를 셀에 추가
                        img.anchor = cell.coordinate
                        ws.add_image(img)

                        # 셀 글자색을 흰색으로 변경
                        cell.font = Font(color="FFFFFF")

            wb.save(temp_filename)
            wb.close()

            return temp_filename

        except Exception as e:
            print("파일 저장 실패", f"파일 저장에 실패했습니다: {e}")
            return None

    def processFileWithHCell(self, source_file):
        target_fname, _ = QFileDialog.getSaveFileName(None, 'Save file as HCell', './', 'HCell files (*.cell)')
        if target_fname:
            try:
                print("한셀 애플리케이션 시작...")
                hcell_start = time.time()
                hcell = Dispatch('HCell.Application')
                hcell.Visible = True
                hcell_end = time.time()
                print(f"한셀 애플리케이션 시작 시간: {hcell_end - hcell_start:.2f}초")

                print("소스 파일 열기...")
                open_start = time.time()
                workbook = hcell.Workbooks.Open(source_file)
                open_end = time.time()
                print(f"소스 파일 열기 시간: {open_end - open_start:.2f}초")

                worksheet = workbook.Sheets(1)

                try:
                    count = 0
                    print("기존 이미지 정보 저장 중...")
                    image_info_start = time.time()
                    images_info = {}
                    for shape in worksheet.Shapes:
                        if shape.Type == 13:
                            images_info[shape.TopLeftCell.Address] = shape
                    image_info_end = time.time()
                    print(f"기존 이미지 정보 저장 시간: {image_info_end - image_info_start:.2f}초")

                    print("새 이미지 배치 시작...")
                    image_place_start = time.time()

                    # 총 셀 수를 계산하고 QProgressDialog 생성
                    total_cells = len(worksheet.UsedRange.Cells)
                    progress_dialog = QProgressDialog("새 이미지 배치 중...", "취소", 0, total_cells)
                    progress_dialog.setWindowTitle("진행 상황")
                    progress_dialog.setWindowModality(Qt.WindowModal)
                    progress_dialog.show()
                    QApplication.processEvents()

                    current_cell = 0
                    for cell in worksheet.UsedRange.Cells:
                        current_cell += 1
                        if cell.Text.endswith(('.png', '.gif')) and os.path.exists(cell.Text):
                            if cell.Address in images_info:
                                existing_shape = images_info[cell.Address]
                                left, top, width, height = existing_shape.Left, existing_shape.Top, existing_shape.Width, existing_shape.Height
                            else:
                                left, top, width, height = cell.Left, cell.Top, cell.Width, cell.Height

                            # AddPicture 호출 수정
                            shape = worksheet.Shapes.AddPicture(Filename=cell.Text, LinkToFile=False, SaveWithDocument=True, Left=left, Top=top, Width=width, Height=height)
                            shape.Placement = 1
                            count += 1
                            print(f"{count}번째 새 이미지가 '{cell.Text}' 경로에서 셀 {cell.Address}에 배치되었습니다. Placement가 1로 설정됨.")
                        
                        # 진행 상황 업데이트
                        progress_dialog.setValue(current_cell)
                        if progress_dialog.wasCanceled():
                            break
                        QApplication.processEvents()

                    image_place_end = time.time()
                    print(f"새 이미지 배치 시간: {image_place_end - image_place_start:.2f}초")

                    print("'Image'로 시작하는 기존 이미지 삭제 중...")
                    image_delete_start = time.time()
                    for shape in list(worksheet.Shapes):
                        if shape.Name.startswith("Image"):
                            shape.Delete()
                    image_delete_end = time.time()
                    print(f"기존 이미지 삭제 시간: {image_delete_end - image_delete_start:.2f}초")

                    # 모든 이미지 크기를 한 개의 셀 크기로 조정, 너비는 2배로 늘림
                    print("이미지 크기 조정 중...")
                    resize_images_start = time.time()
                    cell_width = worksheet.Cells(1, 1).Width
                    cell_height = worksheet.Cells(1, 1).Height
                    for shape in worksheet.Shapes:
                        if shape.Type == 13:  # 13은 이미지 타입
                            shape.Width = cell_width
                            shape.Height = cell_height
                    resize_images_end = time.time()
                    print(f"이미지 크기 조정 시간: {resize_images_end - resize_images_start:.2f}초")

                    print(f"총 {count}개의 새 이미지가 배치되었으며, 'Image'로 시작하는 모든 오래된 이미지가 삭제되었습니다.")
                except Exception as e:
                    print(f"이미지 로딩 및 배치 중 오류 발생: {e}")

                print("워크북 저장 중...")
                save_workbook_start = time.time()
                workbook.SaveAs(target_fname, FileFormat=51)
                workbook.Close(False)
                save_workbook_end = time.time()
                print(f"워크북 저장 시간: {save_workbook_end - save_workbook_start:.2f}초")

                hcell.Quit()

                QMessageBox.information(None, "성공", "파일이 한셀 형식으로 성공적으로 저장되었습니다.")
            except Exception as e:
                print(f"파일 처리 중 오류 발생: {e}")
                QMessageBox.critical(None, "처리 오류", f"파일 처리 중 오류 발생: {e}")
        else:
            print("파일 저장이 취소되었습니다.")

# 여기까지 전자서명 명부 종합하기 관련 클래스



# 입소/중식 확인 관련 클래스
class CheckMealDialog(QDialog):
    def __init__(self, tableWidget, parent=None):
        super(CheckMealDialog, self).__init__(parent)
        self.setWindowTitle('입소/중식 확인')
        self.tableWidget = tableWidget

        layout = QVBoxLayout(self)

        # 열 선택 콤보박스 설정
        self.badgeColumnCombo = QComboBox()
        self.entrySignColumnCombo = QComboBox()
        self.mealNotApplyColumnCombo = QComboBox()
        self.mealUnavailableColumnCombo = QComboBox()

        # 콤보박스에 테이블의 열 제목 추가
        for i in range(self.tableWidget.columnCount()):
            col_name = self.tableWidget.horizontalHeaderItem(i).text()
            self.badgeColumnCombo.addItem(col_name)
            self.entrySignColumnCombo.addItem(col_name)
            self.mealNotApplyColumnCombo.addItem(col_name)
            self.mealUnavailableColumnCombo.addItem(col_name)

        layout.addWidget(QLabel('표찰 열:'))
        layout.addWidget(self.badgeColumnCombo)
        layout.addWidget(QLabel('입소서명 열:'))
        layout.addWidget(self.entrySignColumnCombo)
        layout.addWidget(QLabel('중식 미신청 열:'))
        layout.addWidget(self.mealNotApplyColumnCombo)
        layout.addWidget(QLabel('중식 신청불가 열:'))
        layout.addWidget(self.mealUnavailableColumnCombo)

        # 결과 레이블
        self.resultsLabel = QLabel()
        layout.addWidget(self.resultsLabel)

        # 확인 버튼
        confirmButton = QPushButton('확인', self)
        confirmButton.clicked.connect(self.analyzeData)
        layout.addWidget(confirmButton)

    def analyzeData(self):
        badge_col_index = self.badgeColumnCombo.currentIndex()
        entry_sign_col_index = self.entrySignColumnCombo.currentIndex()
        meal_not_apply_col_index = self.mealNotApplyColumnCombo.currentIndex()
        meal_unavailable_col_index = self.mealUnavailableColumnCombo.currentIndex()

        badge_count = self.countNonEmptyCells(badge_col_index)
        entry_sign_count = self.countNonEmptyCells(entry_sign_col_index)
        meal_not_apply_count = self.countTextCells(badge_col_index, meal_not_apply_col_index, ['미', '신', '청'])
        meal_unavailable_count = self.countTextCells(badge_col_index, meal_unavailable_col_index, ['신', '청', '불', '가'])

        meal_apply_count = badge_count - meal_not_apply_count - meal_unavailable_count

        result_text = (f"표찰 : {badge_count}, 입소서명 : {entry_sign_count}, 중식 신청 : {meal_apply_count}, "
                    f"중식 미신청 : {meal_not_apply_count}, 중식 신청불가 : {meal_unavailable_count}")
        self.resultsLabel.setText(result_text)

    def countNonEmptyCells(self, col_index):
        count = 0
        for row in range(self.tableWidget.rowCount()):
            item = self.tableWidget.item(row, col_index)
            if item and item.text():
                count += 1
        return count

    def countTextCells(self, badge_col_index, col_index, keywords):
        count = 0
        for row in range(self.tableWidget.rowCount()):
            badge_item = self.tableWidget.item(row, badge_col_index)  # 표찰 열 아이템
            text_item = self.tableWidget.item(row, col_index)  # 중식 미신청/신청불가 열 아이템
            if badge_item and badge_item.text() and text_item and any(keyword in text_item.text() for keyword in keywords):
                count += 1
        return count
    

# 인원추가 버튼 관련 클래스
class AddRowDialog(QDialog):
    def __init__(self, parent=None):
        super(AddRowDialog, self).__init__(parent)
        self.setWindowTitle('새 인원 추가')
        layout = QVBoxLayout(self)

        # 날짜 입력
        self.dateEdit = QDateEdit(self)
        self.dateEdit.setCalendarPopup(True)
        self.dateEdit.setDate(QDate.currentDate())
        layout.addWidget(QLabel('날짜:'))
        layout.addWidget(self.dateEdit)

        # 이름 입력
        self.nameLineEdit = QLineEdit(self)
        layout.addWidget(QLabel('성명:'))
        layout.addWidget(self.nameLineEdit)

        # 입력 버튼
        self.addButton = QPushButton('입력', self)
        self.addButton.clicked.connect(self.accept)
        layout.addWidget(self.addButton)

    def getData(self):
        return self.dateEdit.date().toString('yyyy-MM-dd'), self.nameLineEdit.text()


# 여기서부터 표찰/총기번호 클래스
class BadgeGunNumberDialog(QDialog):
    def __init__(self, badgeAndGunNumberData, parent=None):
        super().__init__(parent)
        self.setWindowTitle('표찰/총기번호 조회')
        self.badgeAndGunNumberData = badgeAndGunNumberData
        self.layout = QVBoxLayout(self)

        self.badgeLabel = QLabel("표찰:")
        self.badgeLineEdit = QLineEdit(self)
        self.queryButton = QPushButton("조회", self)
        self.queryButton.clicked.connect(self.queryBadgeNumber)

        self.gunNumberLabel = QLabel("총기번호:")
        self.gunNumberLineEdit = QLineEdit(self)
        self.gunNumberLineEdit.setReadOnly(True)

        self.layout.addWidget(self.badgeLabel)
        self.layout.addWidget(self.badgeLineEdit)
        self.layout.addWidget(self.queryButton)
        self.layout.addWidget(self.gunNumberLabel)
        self.layout.addWidget(self.gunNumberLineEdit)

        self.badgeNumber = None  # 결과를 저장할 속성
        self.gunNumber = None    # 결과를 저장할 속성        

    def queryBadgeNumber(self):
        if self.badgeAndGunNumberData is None:
            QMessageBox.warning(self, "경고", "총기/표찰번호 파일이 선택되지 않았습니다.")
            return
        
        badgeNumber = self.badgeLineEdit.text()
        gunNumber = ""
        for row in self.badgeAndGunNumberData:
            if badgeNumber in row:
                gunNumber = next((x for x in row if x != badgeNumber), None)
                if gunNumber:
                    self.gunNumberLineEdit.setText(gunNumber)
                    self.badgeNumber = badgeNumber  # 결과 저장
                    self.gunNumber = gunNumber  # 결과 저장
                    self.accept()
                    return
        QMessageBox.warning(self, "찾을 수 없음", "해당하는 표찰 번호가 없습니다.")

    # QDialog의 accept 메서드 오버라이드
    def accept(self):
        if self.badgeNumber and self.gunNumber:
            self.done(QDialog.Accepted)
        else:
            super().accept()

# 여기까지 표찰/총기번호 클래스


# 여기부터 계좌번호 입력을 위한 별도의 클래스
class AccountNumberDialog(QDialog):
    def __init__(self, parent=None):
        super(AccountNumberDialog, self).__init__(parent)
        self.setWindowTitle('계좌번호 입력')
        self.setFixedSize(1300, 720)  # 다이얼로그의 크기 설정
        
        layout = QVBoxLayout(self)
        
        # 입력 레이블 설정
        self.accountNumberLabel = QLabel(self)
        self.accountNumberLabel.setAlignment(Qt.AlignCenter)
        # 레이블의 높이를 줄이고, 글씨 크기를 줄입니다.
        self.accountNumberLabel.setStyleSheet("background-color: white; color: black; font-size: 90px; height: 96px; border: 1px solid black;")
        layout.addWidget(self.accountNumberLabel)
        
        # 키패드 레이아웃 설정
        keypadLayout = QGridLayout()
        
        # 버튼 레이블과 위치를 정의합니다.
        buttons = {
            '1': (1, 0), '2': (1, 1), '3': (1, 2),
            '4': (2, 0), '5': (2, 1), '6': (2, 2),
            '7': (3, 0), '8': (3, 1), '9': (3, 2),
            '0': (4, 0), 'Delete': (4, 1), 'Enter': (4, 2)
        }
        
        # 버튼 크기 및 스타일 설정
        buttonStyles = "font-size: 24px; min-width: 220px; min-height: 80px;"
        
        # 키패드 버튼 생성 및 레이아웃에 추가
        for btnText, pos in buttons.items():
            button = QPushButton(btnText)
            button.setStyleSheet(buttonStyles)
            button.clicked.connect(lambda _, b=btnText: self.buttonClicked(b))
            keypadLayout.addWidget(button, *pos)
        
        layout.addLayout(keypadLayout)

        self.centerWindow()

    def centerWindow(self):
        screen = QGuiApplication.primaryScreen().availableGeometry()
        size = self.frameGeometry()
        self.move(
            (screen.width() - size.width()) // 2,
            (screen.height() - size.height()) // 2
        )

    def buttonClicked(self, buttonText):
        if buttonText == 'Delete':
            self.accountNumberLabel.setText(self.accountNumberLabel.text()[:-1])
        elif buttonText == 'Enter':
            self.accept()
        else:
            self.accountNumberLabel.setText(self.accountNumberLabel.text() + buttonText)

    def getAccountNumber(self):
        return self.accountNumberLabel.text()
# 여기까지 계좌번호 입력을 위한 별도의 클래스


#여기서부터 은행 선택 창 클래스
class BankSelectionDialog(QDialog):
    def __init__(self, parent=None):
        super(BankSelectionDialog, self).__init__(parent)
        self.setWindowTitle('은행 선택')
        self.setMinimumSize(1200, 800)
        self.layout = QGridLayout(self)
        self.banks = [
            "국민은행 004", "기업은행 003", "NH농협은행 011", "새마을금고 045", "신한은행 088", "수협은행 007", "신협 048", "씨티은행 027", "우리은행 020", "우체국 071", "카카오뱅크 090", "케이뱅크 089", "토스뱅크 092", "하나은행 081", "SC제일은행 023", "경남은행 039", "광주은행 034", "교보증권 261", "다올투자증권 227","대구은행 031", "대신증권 267", "메리츠증권 287", "미래에셋증권 238", "부국증권 290", "부산은행 032", "비엔피파리바은행 061", "산업은행 002",  "삼성증권 240","상호저축은행 050", "수출입은행 008",   "신영증권 291", "신한금융투자 278", "유안타증권 209",  "유진투자증권 280", "이베스트투자증권 265", "전북은행 037",  "제이피모간체이스은행 057", "제주은행 035",  "중국공상은행 062", "지역농축협 012", "케이프투자증권 292", "키움증권 264", "하나증권 270",  "하이투자증권 262", "한국투자증권 243", "한국포스증권 294", "한화투자증권 269", "현대차증권 263",  "산림조합중앙회 064",  "BOA은행 060", "DB금융투자 279", "HSBC은행 054", "KB증권 218", "NH투자증권 247", "SK증권 266"
        ]
        self.initUI()
        self.centerWindow()

    def centerWindow(self):
        screen = QGuiApplication.primaryScreen().availableGeometry()
        size = self.frameGeometry()  # geometry() 대신 frameGeometry()도 정확함
        self.move(
            (screen.width() - size.width()) // 2,
            (screen.height() - size.height()) // 2
        )

    def initUI(self):
        row, col = 0, 0
        for bank in self.banks:
            button = QPushButton(bank)
            button.setFixedSize(160, 50)  # 버튼 크기 설정
            button.clicked.connect(lambda _, b=bank: self.selectBank(b))
            self.layout.addWidget(button, row, col)
            col += 1
            if col >= 7:  # 3개의 버튼이 한 줄에 배치된 경우 다음 줄로 이동
                row += 1
                col = 0

    def selectBank(self, bank):
        self.parent().setBank(bank)  # 부모 위젯의 setBank 메서드 호출
        self.accept()

# 여기까지 은행 선택 창 클래스        


# 여기서 부터 서명 패드에 관련된 클래스
# 서명 패드에 관련된 메서드
class SignaturePadLabel(QLabel):
    def __init__(self, parent=None):
        super(SignaturePadLabel, self).__init__(parent)
        self.drawing = False
        self.lastPoint = QPoint()
        self.image = QImage(1420, 700, QImage.Format_RGB32)
        if self.image.isNull():
            QMessageBox.critical(self, "Image Error", "Failed to create QImage.")        
        self.image.fill(QColor(192, 192, 192))  # 회색으로 초기 채우기
        self.setStyleSheet("background-color: grey; border: 1px solid black;")

    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.drawing = True
            self.lastPoint = event.pos()

    def mouseMoveEvent(self, event):
        if event.buttons() & Qt.LeftButton and self.drawing:
            self.drawLineTo(event.pos())

    def mouseReleaseEvent(self, event):
        if event.button() == Qt.LeftButton and self.drawing:
            self.drawing = False
            self.drawLineTo(event.pos())

    def drawLineTo(self, endPoint):
        painter = QPainter(self.image)
        painter.setPen(QPen(Qt.black, 5, Qt.SolidLine))
        painter.drawLine(self.lastPoint, endPoint)
        self.lastPoint = endPoint
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.drawImage(0, 0, self.image)

    def clearImage(self):
        self.image.fill(QColor(192, 192, 192))
        self.update()
# 여기까지 서명 패드에 관련 별도 클래스


# 데이터 수정을 위한 별도의 클래스
class EditDataDialog(QDialog):
    def __init__(self, parent=None, imageID=None, column_name=None):
        super(EditDataDialog, self).__init__(parent)
        self.setWindowTitle('데이터 수정')
        self.imageID = imageID
        self.columnName = column_name
        self.signatureSavePath = getattr(parent, 'signatureSavePath', None)

        # QImage 객체를 여기서 초기화합니다.
        self.image = QImage(1570, 850, QImage.Format_ARGB32)
        self.image.fill(Qt.white)

        layout = QVBoxLayout(self)

        self.lineEdit = QLineEdit(self)
        layout.addWidget(self.lineEdit)

        # 수정 버튼 메서드
        self.updateButton = QPushButton('수정', self)
        self.updateButton.clicked.connect(self.accept)
        layout.addWidget(self.updateButton)

        # 은행 버튼 메서드
        self.bankButton = QPushButton('은행', self)
        self.bankButton.clicked.connect(self.showBankMenu)
        layout.addWidget(self.bankButton)

        # 계좌번호 버튼 추가
        self.accountNumberButton = QPushButton('계좌번호', self)
        self.accountNumberButton.clicked.connect(self.showAccountNumberDialog)
        layout.addWidget(self.accountNumberButton)

        # 표찰/총기번호 버튼 추가
        self.badgeGunNumberButton = QPushButton('표찰/총기번호', self)
        self.badgeGunNumberButton.clicked.connect(self.showBadgeGunNumberDialog)
        layout.addWidget(self.badgeGunNumberButton)

        # 미신청 버튼 추가
        self.unappliedButton = QPushButton('미신청', self)
        self.unappliedButton.clicked.connect(self.setUnapplied)
        layout.addWidget(self.unappliedButton)

        # 서명 버튼 메서드
        self.signButton = QPushButton('서명', self)
        self.signButton.clicked.connect(self.showSignaturePad)
        layout.addWidget(self.signButton)


    # 은행 선택 메뉴를 표시하는 메서드
    def showBankMenu(self):
        self.bankDialog = BankSelectionDialog(self)
        self.bankDialog.show()

    # 은행 정보를 QLineEdit에 설정하는 메서드
    def setBank(self, bank):
        self.lineEdit.setText(bank)
        self.accept()
        self.parent().activateWindow()  # 부모 위젯에 포커스 설정
        self.parent().raise_()  # 부모 위젯을 최상위로 올림        

    # 계좌번호 다이얼로그를 표시하고 결과를 처리하는 메서드
    def showAccountNumberDialog(self):
        dialog = AccountNumberDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            accountNumber = dialog.getAccountNumber()
            self.lineEdit.setText(accountNumber)  # self.lineEdit을 업데이트합니다.
            self.accept()

    # 표찰/총기번호 입력 창을 표시하는 메서드
    def showBadgeGunNumberDialog(self):
        dialog = BadgeGunNumberDialog(self.parent().badgeAndGunNumberData, self)
        if dialog.exec_() == QDialog.Accepted:
            badgeNumber = dialog.badgeNumber
            gunNumber = dialog.gunNumber
            self.lineEdit.setText(f"Badge: {badgeNumber}, Gun: {gunNumber}")
            self.parent().updateBadgeAndGunNumber(badgeNumber, gunNumber)
            self.reject()  # 데이터 수정 창을 닫음

    # 미신청 관련 메서드
    def setUnapplied(self):
        self.lineEdit.setText('미신청')
        self.accept()


 # 여기서부터 서명 관련 메서드
    # 서명 패드 관련 메서드 
    def showSignaturePad(self):
        # 부모 객체(Main Window)에서 tableWidget에 접근
        main_window = self.parent()
        
        if hasattr(main_window, 'tableWidget'):  # 부모 객체에 tableWidget 속성이 있는지 확인
            # 1. 메모리 사용을 줄이기 위해 테이블 위젯 업데이트를 비활성화
            main_window.tableWidget.setUpdatesEnabled(False)
            main_window.tableWidget.blockSignals(True)

        self.signaturePadDialog = QDialog(self)
        self.signaturePadDialog.setWindowTitle("서명 패드")
        self.signaturePadDialog.setFixedSize(1450, 800)

        # QVBoxLayout 인스턴스 생성
        layout = QVBoxLayout(self.signaturePadDialog)
        
        # SignaturePadLabel 인스턴스 생성 및 layout에 추가
        self.label = SignaturePadLabel(self.signaturePadDialog)
        self.label.setFixedSize(1420, 700)
        layout.addWidget(self.label)

        # "서명 저장" 버튼 생성 및 layout에 추가
        saveButton = QPushButton("서명 저장", self.signaturePadDialog)
        saveButton.clicked.connect(lambda: self.saveSignature(self.signaturePadDialog))
        layout.addWidget(saveButton)

        # "서명 지우기" 버튼 생성 및 layout에 추가
        clearButton = QPushButton("서명 지우기", self.signaturePadDialog)
        clearButton.clicked.connect(self.label.clearImage)  # clearImage는 SignaturePadLabel의 메서드
        layout.addWidget(clearButton)
        
        # layout을 signaturePadDialog의 레이아웃으로 설정
        self.signaturePadDialog.setLayout(layout)

        # 다이얼로그가 종료될 때 호출되는 콜백 함수 정의
        self.signaturePadDialog.finished.connect(self.restoreUI)

        # 화면 중앙 정렬
        self.centerWindowSignature()  

        # 다이얼로그 표시
        self.signaturePadDialog.show()

    def centerWindowSignature(self):
        screen = QGuiApplication.primaryScreen().availableGeometry()
        size = self.signaturePadDialog.frameGeometry()
        self.signaturePadDialog.move(
            (screen.width() - size.width()) // 2,
            (screen.height() - size.height()) // 2
        )    

    def restoreUI(self):
        # 부모 객체(Main Window)에서 tableWidget에 접근
        main_window = self.parent()
        
        if hasattr(main_window, 'tableWidget'):
            # 테이블 위젯 업데이트를 다시 활성화
            main_window.tableWidget.setUpdatesEnabled(True)
            main_window.tableWidget.blockSignals(False)

    def eventFilter(self, source, event):
        if source is self.label:
            if event.type() == QEvent.MouseMove and (event.buttons() & Qt.LeftButton):
                print("MouseMove with LeftButton pressed")  # 디버깅 메시지 추가
                self.drawLineTo(event.pos())
            elif event.type() == QEvent.MouseButtonPress and event.button() == Qt.LeftButton:
                print("MouseButtonPress with LeftButton")  # 디버깅 메시지 추가
                self.drawing = True
                self.lastPoint = event.pos()
            elif event.type() == QEvent.MouseButtonRelease and event.button() == Qt.LeftButton:
                print("MouseButtonRelease with LeftButton")  # 디버깅 메시지 추가
                self.drawing = False
            return True
        return super().eventFilter(source, event)
        
    def drawLineTo(self, endPoint):
        if self.drawing:
            painter = QPainter(self.image)
            painter.setPen(QPen(Qt.black, 5, Qt.SolidLine))
            painter.drawLine(self.lastPoint, endPoint)
            self.lastPoint = endPoint
            painter.end()
            self.updateCanvas()

    def updateCanvas(self):
        self.label.setPixmap(QPixmap.fromImage(self.image))
    
    def clearSignature(self):
        self.image.fill(Qt.white)
        self.updateCanvas()

    def saveSignature(self, dialog):
        # scene 관련 부분을 제거하고 self.image를 직접 사용하여 저장합니다.
        if not self.signatureSavePath:
            QMessageBox.warning(self, "경로 오류", "서명 이미지를 저장할 경로가 설정되지 않았습니다.")
            return

        # 파일 이름 및 경로 설정
        fileName = f"{self.imageID}_{self.columnName}_signature.png"  
        filePath = os.path.join(self.signatureSavePath, fileName)

        # 저장 경로에 디렉토리가 존재하는지 확인하고, 없으면 생성
        os.makedirs(self.signatureSavePath, exist_ok=True)

        # 디버깅: imageID와 columnName 확인
        print("Debug Info:")
        print(f"imageID: {self.imageID}")
        print(f"columnName: {self.columnName}")
        print(f"filePath: {filePath}")

        # QImage 객체를 QPixmap으로 변환
        pixmap = QPixmap.fromImage(self.label.image)
        # QPixmap을 Pillow Image로 변환
        if pixmap.isNull():
            QMessageBox.warning(self, "변환 오류", "이미지를 변환하는 데 실패했습니다.")
            return
        else:
            buffer = QBuffer()
            buffer.open(QBuffer.ReadWrite)
            pixmap.save(buffer, "PNG")
            pil_img = PilImage.open(io.BytesIO(buffer.data()))

        # 이미지 크기 조정
        try:
            # 셀 크기로 이미지 크기 조정
            img_resized = pil_img.resize((60, 18), PilImage.LANCZOS)
            # PNG 형식으로 이미지 저장
            img_resized.save(filePath, "PNG")
            QMessageBox.information(self, "저장됨", "서명이 저장되었습니다.")
            self.lineEdit.setText(filePath)
        except Exception as e:
            QMessageBox.warning(self, "저장 오류", f"서명 이미지를 저장하는데 실패했습니다: {str(e)}")

        self.restoreUI()  # UI를 복구하는 함수 호출

        # 저장된 파일 경로를 lineEdit에 입력
        self.lineEdit.setText(filePath)
        self.accept()

        # 다이얼로그 닫기
        dialog.accept()

   # 여기까지 서명 관련 메서드


#ID열 선택 창 클래스
class ColumnSelectorDialog(QDialog):
    def __init__(self, columnNames, parent=None):
        super().__init__(parent)
        self.setWindowTitle('열 선택')
        self.layout = QVBoxLayout(self)

        self.columnComboboxes = {}
        labels = [
            ('훈련시작일차 열 선택', '훈련시작일차'),
            ('훈련종료일차 열 선택', '훈련종료일차'),
            ('성명 열 선택', '성명'),
            ('생년월일 열 선택', '생년월일'),
            ('훈련유형 열 선택', '훈련유형'),
            ('개인차수 열 선택', '개인차수'),
            ('총 계획시간 열 선택', '총계획시간'),
            ('당일 총기번호 열 선택', ''), # 자동선택 안함
            ('1일차 훈련시간 열 선택', '1일차 훈련시간'),
            ('2일차 훈련시간 열 선택', '2일차 훈련시간'),
            ('3일차 훈련시간 열 선택', '3일차 훈련시간'),
            ('4일차 훈련시간 열 선택', '4일차 훈련시간'),
            ('5일차 훈련시간 열 선택', '5일차 훈련시간')
        ]

        for label, keyword in labels:
            rowLayout = QHBoxLayout()
            labelWidget = QLabel(label)
            comboBox = QComboBox()
            comboBox.addItems(columnNames)

            # 자동 선택 로직
            if keyword:
                index = next((i for i, name in enumerate(columnNames) if keyword in name), -1)
                comboBox.setCurrentIndex(index)
            else:
                comboBox.setCurrentIndex(-1)  # 자동선택 안 함

            self.columnComboboxes[label] = comboBox
            rowLayout.addWidget(labelWidget)
            rowLayout.addWidget(comboBox)
            self.layout.addLayout(rowLayout)

        # 확인 버튼 추가
        self.okButton = QPushButton('확인', self)
        self.okButton.clicked.connect(self.accept)
        self.layout.addWidget(self.okButton)

        # 다이얼로그를 항상 위로 설정
        self.setWindowFlag(Qt.WindowStaysOnTopHint)

        # 다이얼로그를 가장 위로 올리고 포커스 맞추기
        self.raise_()
        self.activateWindow()

    # ID열 콤보박스 선택 관련 메서드
    def getColumnSelections(self):
        return {label: comboBox.currentText() for label, comboBox in self.columnComboboxes.items() if comboBox.currentIndex() != -1}



# 메인 애플리케이션 클래스
class SignatureApp(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()  # 사용자 인터페이스 초기화

        # 속성을 초기화
        self.signatureSavePath = None
        self.badgeAndGunNumberData = None
        self.gunNumberColumnName = None 

        # QTableWidget의 item이 변경될 때마다 onItemChanged 메서드를 호출하도록 연결
        #self.tableWidget.itemChanged.connect(self.onItemChanged)
        
        self.searchResults = []  # 검색 결과를 저장할 리스트
        self.currentSearchIndex = -1  # 현재 검색 결과 인덱스

    # GUI 메인 메서드
    def initUI(self):
        self.setWindowTitle('지역예비군훈련 전자결산 프로그램')  # 윈도우 타이틀 설정
        self.setGeometry(100, 100, 1800, 400)  # 윈도우 크기와 위치 설정
        mainLayout = QVBoxLayout()  # 메인 레이아웃 생성

# 여기서부터 컨트롤 프레임      
          

        self.controlFrame = QFrame(self)  # 버튼과 레이블을 포함할 프레임 생성
        self.controlFrame.setFrameShape(QFrame.StyledPanel)  # 프레임 스타일 설정
        self.controlFrame.setFixedHeight(200)  # 프레임 높이 설정

        mainLayout.addWidget(self.controlFrame)  # 프레임을 메인 레이아웃에 추

        # 레이블 초기화
        self.fileNameLabel = QLabel("전자서명 명부 파일 : 없음", self.controlFrame)
        self.selectedBadgeAndGunLabel = QLabel("총기/표찰번호 파일: 없음", self.controlFrame)
        self.signatureSavePathLabel = QLabel("서명 이미지 폴더: 없음", self.controlFrame)

        # 1번 프레임 추가 및 설정
        frame1 = QFrame(self.controlFrame)
        frame1.setFrameShape(QFrame.StyledPanel)
        frame1.setGeometry(10, 10, 250, 180)  # 위치와 크기 설정

        self.selectFileButton = QPushButton('전자서명 명부 불러오기', frame1)
        self.selectFileButton.setGeometry(10, 10, 230, 30)
        self.selectFileButton.clicked.connect(self.selectFile)

        self.loadBadgeAndGunButton = QPushButton('표찰/총기번호 불러오기', frame1)
        self.loadBadgeAndGunButton.setGeometry(10, 50, 230, 30)
        self.loadBadgeAndGunButton.clicked.connect(self.loadBadgeAndGunNumber)

        self.fileNameLabel.setParent(frame1)
        self.fileNameLabel.setGeometry(10, 90, 230, 20)

        self.selectedBadgeAndGunLabel.setParent(frame1)
        self.selectedBadgeAndGunLabel.setGeometry(10, 110, 230, 20)

        self.signatureSavePathLabel.setParent(frame1)
        self.signatureSavePathLabel.setGeometry(10, 130, 230, 20)

        # 2번 프레임 추가 및 설정
        frame2 = QFrame(self.controlFrame)
        frame2.setFrameShape(QFrame.StyledPanel)
        frame2.setGeometry(270, 10, 300, 180)  # 위치와 크기 설정

        self.saveFileButton = QPushButton('전자서명 명부 저장하기', frame2)
        self.saveFileButton.setGeometry(10, 10, 280, 30)
        self.saveFileButton.clicked.connect(self.saveFile)

        self.saveFileWithoutImagesButton = QPushButton('전자서명 명부 저장하기(이미지 미포함)', frame2)
        self.saveFileWithoutImagesButton.setGeometry(10, 50, 280, 30)
        self.saveFileWithoutImagesButton.clicked.connect(self.saveFileWithoutImages)

        self.mergeSignatureFilesButton = QPushButton('전자서명 명부 종합하기', frame2)
        self.mergeSignatureFilesButton.setGeometry(10, 90, 280, 30)
        self.mergeSignatureFilesButton.clicked.connect(self.openMergeDialog)

        self.checkMealTransportButton = QPushButton('훈련 참가비 확인(CMS)', frame2)
        self.checkMealTransportButton.setGeometry(10, 130, 280, 30)
        self.checkMealTransportButton.clicked.connect(self.openCheckMealTransportDialog)

        # 3번 프레임 추가 및 설정
        frame3 = QFrame(self.controlFrame)
        frame3.setFrameShape(QFrame.StyledPanel)
        frame3.setGeometry(580, 10, 250, 180)  # 위치와 크기 설정

        self.saveTrainingSummaryButton = QPushButton('훈련결산 저장하기', frame3)
        self.saveTrainingSummaryButton.setGeometry(10, 10, 230, 30)
        self.saveTrainingSummaryButton.clicked.connect(self.openSaveTrainingSummaryDialog)

        self.summarizeTrainingButton = QPushButton('훈련결산 종합하기(동원2형)', frame3)
        self.summarizeTrainingButton.setGeometry(10, 50, 230, 30)
        self.summarizeTrainingButton.clicked.connect(self.openTrainingSummaryDialog)

        # 4번 프레임 추가 및 설정
        frame4 = QFrame(self.controlFrame)
        frame4.setFrameShape(QFrame.StyledPanel)
        frame4.setGeometry(840, 10, 250, 180)  # 위치와 크기 설정

        self.checkMealButton = QPushButton('입소/중식 확인', frame4)
        self.checkMealButton.setGeometry(10, 10, 230, 30)
        self.checkMealButton.clicked.connect(self.openCheckMealDialog)

        # 전자서명 명부 서식 생성 버튼 추가
        self.createSignatureFormButton = QPushButton('(원본)전자서명 명부 서식 생성', frame4)
        self.createSignatureFormButton.setGeometry(10, 50, 230, 30)
        self.createSignatureFormButton.clicked.connect(self.createSignatureForm)        

        self.setLayout(mainLayout)  # 위젯에 메인 레이아웃 설정

# 여기까지 컨트롤 프레임
        
# 여기서부터 중간 프레임
        
        # 중간 프레임 설정
        self.middleFrame = QFrame(self)  # 중간 프레임 생성
        self.middleFrame.setFrameShape(QFrame.StyledPanel)  # 프레임 스타일 설정
        self.middleLayout = QHBoxLayout(self.middleFrame)  # 중간 프레임 내 레이아웃 생성
        mainLayout.addWidget(self.middleFrame)  # 중간 프레임을 메인 레이아웃에 추가

        # 검색 관련 위젯을 위한 레이아웃
        searchLayout = QHBoxLayout()
        self.searchLineEdit = QLineEdit(self.middleFrame)
        self.searchLineEdit.setFixedWidth(200)  # 검색 창의 너비 조정
        self.searchButton = QPushButton('검색', self.middleFrame)
        self.searchButton.setFixedWidth(100)
        self.searchButton.clicked.connect(self.searchInTable)
        searchLayout.addWidget(self.searchLineEdit)
        searchLayout.addWidget(self.searchButton)

        # 전체 중간 레이아웃에 검색 레이아웃 추가
        self.middleLayout.addLayout(searchLayout)

        # 열 관리 창
        self.manageColumnsButton = QPushButton('열 관리', self.middleFrame)
        self.manageColumnsButton.setFixedWidth(100)
        self.manageColumnsButton.clicked.connect(self.manageColumns)
        searchLayout.addWidget(self.manageColumnsButton)

        # 인원추가 버튼 추가
        self.addRowButton = QPushButton('인원추가', self.middleFrame)
        searchLayout.addWidget(self.addRowButton)
        self.addRowButton.clicked.connect(self.addNewRow)        

        # 데이터 조회 레이블 추가
        self.dataCountLabel = QLabel("데이터 조회 : 0", self.middleFrame)
        searchLayout.addWidget(self.dataCountLabel)

        # 중간 레이아웃에 스트레치 추가하여 검색 위젯과 메뉴 숨기기/보이기 버튼 사이 공간 확보
        self.middleLayout.addStretch(1)  # 동적으로 공간 추가

        # 메뉴 숨기기/보이기 버튼 추가, 이 버튼은 화면의 우측 끝에 위치하게 됩니다.
        self.toggleFrameButton = QPushButton('메뉴 숨기기/보이기', self.middleFrame)
        self.toggleFrameButton.clicked.connect(self.toggleFrameVisibility)
        self.toggleFrameButton.setFixedSize(150, 30)
        self.middleLayout.addWidget(self.toggleFrameButton)  # 중간 레이아웃에 버튼 직접 추가

# 여기까지 중간 프레임

# 여기서부터 프렘임 외 내용

        self.tableWidget = QTableWidget()  # 데이터를 표시할 QTableWidget 생성
        mainLayout.addWidget(self.tableWidget)  # 테이블 위젯을 메인 레이아웃에 추가

        self.tableWidget.itemChanged.connect(self.updateRowCount)  # 테이블 변경 시 레이블 업데이트
        self.tableWidget.model().rowsInserted.connect(self.updateRowCount)
        self.tableWidget.model().rowsRemoved.connect(self.updateRowCount)

        self.setLayout(mainLayout)  # 위젯에 메인 레이아웃 설정

        # 위젯 더블클릭시 호출하는 내용
        self.tableWidget.itemDoubleClicked.connect(self.editItem)

    # 데이터 수정 창 관련 메서드
    def editItem(self, item):
        try:
            # 클릭한 항목의 열 인덱스를 가져옵니다.
            column_index = item.column()
            # 열 이름을 헤더에서 추출합니다.
            column_name = self.tableWidget.horizontalHeaderItem(column_index).text()

            # 기존 로그 출력 및 기본 설정 코드 유지
            print("badgeAndGunNumberData:", self.badgeAndGunNumberData)
            print("gunNumberColumnName:", self.gunNumberColumnName)

            row = item.row()  # 사용자가 선택한 셀의 행

            # 총기번호 열의 인덱스를 찾음
            gunNumberColumnIndex = self.findColumnIndex(self.gunNumberColumnName)

            print("gunNumberColumnIndex:", gunNumberColumnIndex)

            # 선택한 셀의 행에 해당하는 ID 열의 데이터를 가져옴
            idItem = self.tableWidget.item(row, 0)
            if idItem is None or not idItem.text():
                QMessageBox.warning(self, "오류", "유효한 imageID를 찾을 수 없습니다.")
                return

            imageID = idItem.text()

            # 더블클릭한 셀 위치를 저장
            self.currentRow = item.row()
            self.currentCol = item.column()

            # 기존 EditDataDialog 실행 코드 유지
            dialog = EditDataDialog(self, imageID, column_name)
            dialog.lineEdit.setText(item.text())
            if dialog.exec_() == QDialog.Accepted:
                updatedText = dialog.lineEdit.text()
                item.setText(updatedText)  # 입력 필드의 내용으로 셀의 텍스트를 업데이트
        except Exception as e:
            QMessageBox.critical(self, "오류 발생", f"데이터 수정 중 예외가 발생했습니다: {e}")

  # 여기서부터 표찰/총기번호 설정 관련 메서드

    # BadgeGunNumberDialog에서 반환된 표찰번호와 총기번호를 처리하는 새 메서드
    def updateBadgeAndGunNumber(self, badgeNumber, gunNumber):
        print(f"updateBadgeAndGunNumber 메서드가 호출되었습니다. 표찰번호: {badgeNumber}, 총기번호: {gunNumber}")  # 로그 출력 추가
        if self.currentRow is not None and self.currentCol is not None:
            # 더블클릭한 셀에 표찰번호를 업데이트
            self.tableWidget.setItem(self.currentRow, self.currentCol, QTableWidgetItem(badgeNumber))
            # 총기번호 열 인덱스를 찾고 해당 열에 총기번호를 업데이트
            gunNumberColumnIndex = self.findColumnIndex(self.gunNumberColumnName)
            if gunNumberColumnIndex != -1:
                self.tableWidget.setItem(self.currentRow, gunNumberColumnIndex, QTableWidgetItem(gunNumber))

    # 주어진 열 이름에 해당하는 테이블 위젯의 열 인덱스를 찾는 메서드
    def findColumnIndex(self, columnName):
        for i in range(self.tableWidget.columnCount()):
            if self.tableWidget.horizontalHeaderItem(i).text() == columnName:
                return i
        return -1
    
 # 여기까지 표찰/총기번호 설정 관련 메서드

# 여기까지 프레임 외의 메서드


# 여기서부터 컨트롤 프레임 관련 메서드

    # 상단 프레임 토글(보이고 안보이고)하는 메서드
    def toggleFrameVisibility(self):
        self.controlFrame.setVisible(not self.controlFrame.isVisible())  # 프레임의 가시성 토글

   # 여기서부터 json 데이터 저장 관련 메서드

    # 테이블 위젯에서 데이터를 추출하는 메서드
    def getTableData(self):
        data = []
        for row in range(self.tableWidget.rowCount()):
            row_data = {}
            for col in range(self.tableWidget.columnCount()):
                header = self.tableWidget.horizontalHeaderItem(col).text()  # 열 헤더의 텍스트를 가져옴
                cell = self.tableWidget.item(row, col)
                row_data[header] = cell.text() if cell else ""  # 셀 데이터를 추출하거나 셀이 없다면 빈 문자열 사용
            data.append(row_data)
        return data

    # 전자서명 명부 종합하기 버튼의 클릭 이벤트에 연결될 메서드
    def openMergeDialog(self):
        # 서명 이미지 경로 설정 다이얼로그 표시
        savePath = QFileDialog.getExistingDirectory(self, "서명 이미지 저장 경로 선택")
        if savePath:
            self.signatureSavePath = savePath  # 선택된 경로 저장
            self.signatureSavePathLabel.setText(f"서명 이미지 폴더: {savePath}")  # 레이블에 경로 표시
        else:
            QMessageBox.warning(self, "경로 선택 오류", "서명 이미지를 저장할 경로가 선택되지 않았습니다.")
            return  # 경로 선택이 취소되면 함수 종료

        dialog = MergeFilesDialog(self, self.signatureSavePath)
        dialog.exec_()

    # 입소/중식 확인 버튼 호출 메서드
    def openCheckMealDialog(self):
        dialog = CheckMealDialog(self.tableWidget, self)
        dialog.exec_()

    # 전자서명 명부 불러오기 위한 선택 메서드
    def selectFile(self):

        # 2단계: 사용자로부터 서명 이미지 저장 경로 선택받기
        savePath = QFileDialog.getExistingDirectory(self, "서명 이미지 저장 경로 선택")
        if not savePath:
            # 사용자가 경로 선택을 취소했다면, 이후의 로직을 실행하지 않음
            QMessageBox.warning(self, "경로 선택 오류", "서명 이미지를 저장할 경로가 선택되지 않았습니다.")
            return
        else:
            self.signatureSavePath = savePath  # 선택된 경로 저장
            self.signatureSavePathLabel.setText(f"서명 이미지 폴더: {savePath}")  # 레이블에 경로 표시

        # 3단계: 사용자로부터 파일 선택받기
        fname, _ = QFileDialog.getOpenFileName(self, 'Open file', './', 'hcell files (*.cell)')
        if not fname:
            QMessageBox.warning(self, "파일 선택 오류", "파일이 선택되지 않았습니다.")
            return
        self.fileNameLabel.setText(f"선택된 파일: {fname}")  # 선택된 파일 이름을 레이블에 표시

        # 4단계: 선택된 파일 처리
        try:
            # os.startfile을 사용하여 파일 열기
            os.startfile(fname)

            # 테이블 형태로 데이터를 입력할 수 있는 창 생성
            tableDialog = QDialog(self)
            tableDialog.setWindowTitle("데이터 확인 및 입력")
            layout = QVBoxLayout(tableDialog)

            # 설명 라벨 추가
            instructionsLabel = QLabel("엑셀에서 데이터를 복사하여 아래 텍스트 박스에 붙여넣기 하십시오.")
            layout.addWidget(instructionsLabel)

            # 텍스트 에디터 추가
            textEdit = QPlainTextEdit()
            layout.addWidget(textEdit)

            # 텍스트가 변경될 때마다 호출되는 함수 정의
            def validateText():
                copied_data = textEdit.toPlainText().strip()
                table_data = [row.split('\t') for row in copied_data.split('\n')]

                # 1. 각 줄의 열 수를 제한하여 70열을 넘는 데이터를 제거
                max_columns = 70
                table_data = [row[:max_columns] for row in table_data]

                # 2. 줄 수 제한 (2000줄까지만 허용)
                max_rows = 2000
                if len(table_data) > max_rows:
                    table_data = table_data[:max_rows]

                # 수정된 데이터를 다시 텍스트로 변환하여 텍스트 에디터에 입력
                corrected_text = '\n'.join(['\t'.join(row) for row in table_data])
                
                # 텍스트가 다르면 수정된 텍스트로 교체
                if corrected_text != copied_data:
                    # 이벤트 일시 중지
                    textEdit.blockSignals(True)
                    textEdit.setPlainText(corrected_text)
                    textEdit.blockSignals(False)

            # 텍스트가 변경될 때마다 validateText 함수를 호출
            textEdit.textChanged.connect(validateText)

            # 확인 버튼 추가
            confirmButton = QPushButton("확인", tableDialog)
            confirmButton.clicked.connect(lambda: self.processCopiedData(textEdit, tableDialog))
            layout.addWidget(confirmButton)

            # 다이얼로그 실행
            tableDialog.exec_()

        except Exception as e:
            QMessageBox.critical(self, "오류 발생", f"파일을 불러오는 데 실패했습니다: {e}")

    def processCopiedData(self, textEdit, tableDialog):
        copied_data = textEdit.toPlainText().strip()
        table_data = [row.split('\t') for row in copied_data.split('\n')]

        columnNames = table_data[0]
        columnDialog = ColumnSelectorDialog(columnNames, self)
        if columnDialog.exec_() == QDialog.Accepted:
            columnSelections = columnDialog.getColumnSelections()
            self.gunNumberColumnName = columnSelections.get('당일 총기번호 열 선택')
            # ID 열이 있는지 여부를 체크하여 전달
            idColumnIndex = columnNames.index('ID') if 'ID' in columnNames else -1
            self.displayDataInTable(table_data, columnSelections, idColumnIndex)

            QMessageBox.information(self, 'Success', '파일을 불러오는데 성공했습니다')

        tableDialog.accept()

    # 전자서명 명부로 불러온 파일을 화면에 보여주는 메서드            
    def displayDataInTable(self, table_data, columnSelections, idColumnIndex):
        # 기존 데이터의 열 제목 읽기
        headers = table_data[0]
        print(f"Headers: {headers}")  # 디버깅 출력

        # 테이블 업데이트 비활성화
        self.tableWidget.setUpdatesEnabled(False)        

        # 기존에 'ID' 열이 있는지 확인
        id_present = 'ID' in headers
        if id_present:
            idColumnIndex = headers.index('ID')
            headers.pop(idColumnIndex)  # 기존 'ID' 열 제목 제거
        headers = ['ID'] + headers  # 새로운 'ID' 열을 제목에 추가
        self.tableWidget.setColumnCount(len(headers))  # 열 수 조정

        self.tableWidget.setHorizontalHeaderLabels(headers)  # 열 제목 설정
        # 데이터 로우 설정 (두 번째 행부터 시작)
        self.tableWidget.setRowCount(len(table_data) - 1)  # 첫 번째 행(제목)을 제외한 나머지 행 수

        # 선택된 열의 인덱스를 찾거나 빈 문자열로 설정합니다.
        def get_corrected_index(col_name):
            original_index = headers.index(col_name) if col_name in headers else -1
            if id_present and original_index >= idColumnIndex:
                return original_index + 1  # ID 열을 제거한 후 인덱스 조정
            return original_index

        trainingStartDayIndex = get_corrected_index(columnSelections.get('훈련시작일차 열 선택'))
        trainingEndDayIndex = get_corrected_index(columnSelections.get('훈련종료일차 열 선택'))
        nameIndex = get_corrected_index(columnSelections.get('성명 열 선택'))
        birthDateIndex = get_corrected_index(columnSelections.get('생년월일 열 선택'))
        trainingTypeIndex = get_corrected_index(columnSelections.get('훈련유형 열 선택'))
        individualSequenceIndex = get_corrected_index(columnSelections.get('개인차수 열 선택'))
        totalPlanTimeIndex = get_corrected_index(columnSelections.get('총 계획시간 열 선택'))

        dailyTrainingTimeIndexes = [
            get_corrected_index(columnSelections.get('1일차 훈련시간 열 선택')),
            get_corrected_index(columnSelections.get('2일차 훈련시간 열 선택')),
            get_corrected_index(columnSelections.get('3일차 훈련시간 열 선택')),
            get_corrected_index(columnSelections.get('4일차 훈련시간 열 선택')),
            get_corrected_index(columnSelections.get('5일차 훈련시간 열 선택'))
        ]

        # 디버깅 출력
        print(f"Column Selections: {columnSelections}")
        print(f"Training Start Day Index: {trainingStartDayIndex} (열 이름: {headers[trainingStartDayIndex] if trainingStartDayIndex != -1 else 'N/A'})")
        print(f"Training End Day Index: {trainingEndDayIndex} (열 이름: {headers[trainingEndDayIndex] if trainingEndDayIndex != -1 else 'N/A'})")
        print(f"Name Index: {nameIndex} (열 이름: {headers[nameIndex] if nameIndex != -1 else 'N/A'})")
        print(f"Birth Date Index: {birthDateIndex} (열 이름: {headers[birthDateIndex] if birthDateIndex != -1 else 'N/A'})")
        print(f"Training Type Index: {trainingTypeIndex} (열 이름: {headers[trainingTypeIndex] if trainingTypeIndex != -1 else 'N/A'})")
        print(f"Individual Sequence Index: {individualSequenceIndex} (열 이름: {headers[individualSequenceIndex] if individualSequenceIndex != -1 else 'N/A'})")
        print(f"Total Plan Time Index: {totalPlanTimeIndex} (열 이름: {headers[totalPlanTimeIndex] if totalPlanTimeIndex != -1 else 'N/A'})")

        for i, index in enumerate(dailyTrainingTimeIndexes, start=1):
            print(f"Day {i} Training Time Index: {index} (열 이름: {headers[index] if index != -1 else 'N/A'})")

        # 모든 행에 대해 새로운 'ID' 값을 계산하고 나머지 데이터와 함께 테이블에 설정
        for row in range(1, len(table_data)):
            # 새로운 'ID' 값을 구성하기 위한 데이터 추출
            idValues = []
            missingColumns = []  # 데이터가 없는 열을 기록하는 리스트
            
            # try-except 블록으로 인덱스 에러 방지
            try:
                # 각 열에 대해 데이터를 확인하고, 없으면 경고창에 알림을 추가합니다.
                for colName, colIndex in zip(['성명', '생년월일', '훈련유형', '개인차수'], 
                                            [nameIndex - 1, birthDateIndex - 1, trainingTypeIndex - 1, individualSequenceIndex - 1]):
                    if colIndex != -1 and colIndex < len(table_data[row]):  # 인덱스 유효성 검사
                        cellValue = table_data[row][colIndex]
                        if cellValue is not None:
                            if isinstance(cellValue, float) and cellValue.is_integer():
                                cellValue = int(cellValue)
                            cellValue = str(cellValue)
                            idValues.append(cellValue)  # ID 구성 값 추가
                        else:
                            missingColumns.append(colName)  # 값이 없는 열 기록
                    else:
                        missingColumns.append(colName)  # 인덱스 범위를 벗어난 열 기록
                
                # 데이터가 없는 열이 있을 경우 경고 메시지 표시
                if missingColumns:
                    missingColsString = ", ".join(missingColumns)
                    QMessageBox.warning(self, "데이터 오류", f"행 {row}에서 다음 열에 데이터가 없습니다: {missingColsString}")
                
                # ID 값 생성
                idValue = "_".join(idValues) if idValues else ""

                if not idValue:  # ID 열이 없는 경우 해당 행을 건너뛰기
                    continue

                # 테이블에 새로운 'ID' 열 값 설정
                self.tableWidget.setItem(row - 1, 0, QTableWidgetItem(idValue))

            except IndexError as e:
                # 예외 처리: 인덱스 에러 발생 시 경고창만 띄우고 프로그램 종료 방지
                QMessageBox.critical(self, "인덱스 오류", f"행 {row}에서 인덱스 오류가 발생했습니다: {e}")
                continue  # 다음 행으로 넘어가서 계속 실행

            adjustedColumnIndex = 1  # 새로운 'ID' 열 이후부터 시작
            # 'ID' 열 제외하고 값을 읽어옴
            for col in range(0, len(table_data[0])):
                if col == idColumnIndex:  # 'ID' 열이면 건너뛰기
                    continue

                value = table_data[row][col] if col < len(table_data[row]) else None  # 셀 값 가져오기
                if value is None:  # 값이 None이면 빈 문자열로 처리
                    displayValue = ""
                elif isinstance(value, float) and value.is_integer():
                    value = int(value)  # 부동소수점 수가 정수면 정수 형태로 변환
                    displayValue = str(value)
                else:
                    displayValue = str(value)

                # 빈 값, '0', '00', '000' 조건 검사
                if displayValue in ['', '0', '00', '000']:
                    displayValue = ""  # 이 조건들을 만족하면 테이블에 아무것도 표시하지 않음

                item = QTableWidgetItem(displayValue)
                self.tableWidget.setItem(row - 1, adjustedColumnIndex, item)  # 테이블에 데이터 설정
                adjustedColumnIndex += 1  # 열 인덱스 조정

            # 총 계획시간, 훈련시작일차, 훈련종료일차 인덱스를 조정하여 가져오기
            adjustedTotalPlanTimeIndex = totalPlanTimeIndex - 1 if totalPlanTimeIndex > 0 else totalPlanTimeIndex
            adjustedTrainingStartDayIndex = trainingStartDayIndex - 1 if trainingStartDayIndex > 0 else trainingStartDayIndex
            adjustedTrainingEndDayIndex = trainingEndDayIndex - 1 if trainingEndDayIndex > 0 else trainingEndDayIndex

            # 총 계획시간, 훈련시작일차, 훈련종료일차 읽기
            totalPlanHoursValue = table_data[row][adjustedTotalPlanTimeIndex] if adjustedTotalPlanTimeIndex != -1 and adjustedTotalPlanTimeIndex < len(table_data[row]) else '0'
            print(f"행 {row-1}: 총 계획시간 값: {totalPlanHoursValue}")

            if totalPlanHoursValue in [None, '']:
                totalPlanHours = 0  # None이거나 빈 문자열일 경우 0으로 설정
            else:
                try:
                    totalPlanHours = float(totalPlanHoursValue)
                except ValueError:
                    print(f"행 {row-1}: 총 계획시간 값을 float로 변환하는 중 오류: {totalPlanHoursValue}")
                    totalPlanHours = 0  # 기본값 설정 또는 다른 처리
            print(f"행 {row-1}: 총 계획시간: {totalPlanHours}")

            # 훈련시작일차와 훈련종료일차가 숫자인지 확인
            startDayValue = table_data[row][adjustedTrainingStartDayIndex] if adjustedTrainingStartDayIndex != -1 and adjustedTrainingStartDayIndex < len(table_data[row]) else '1'
            endDayValue = table_data[row][adjustedTrainingEndDayIndex] if adjustedTrainingEndDayIndex != -1 and adjustedTrainingEndDayIndex < len(table_data[row]) else '5'
            print(f"행 {row-1}: 시작일차 원본 값: {startDayValue}, 종료일차 원본 값: {endDayValue}")

            try:
                startDay = int(startDayValue)
            except ValueError:
                print(f"행 {row-1}: 시작일차 값을 int로 변환하는 중 오류: {startDayValue}")
                startDay = 1  # 기본값으로 설정하거나 다른 처리
            print(f"행 {row-1}: 시작일차: {startDay}")

            # 종료일차 값에서 숫자만 추출
            import re
            endDayValueNumbers = re.findall(r'\d+', endDayValue)
            endDayValue = endDayValueNumbers[0] if endDayValueNumbers else '5'

            try:
                endDay = int(endDayValue)
            except ValueError:
                print(f"행 {row-1}: 종료일차 값을 int로 변환하는 중 오류: {endDayValue}")
                endDay = 5  # 기본값으로 설정하거나 다른 처리
            print(f"행 {row-1}: 종료일차: {endDay}")

            # 디버깅 출력
            print(f"행 {row-1}: 총 계획시간: {totalPlanHours}, 시작일차: {startDay}, 종료일차: {endDay}")

            # 수정된 일차별 훈련시간 계산 로직
            if totalPlanHours > 0:
                # 계획된 훈련 기간에 맞게 시간 분배
                allocatedHours = [0] * 5  # 각 일차별로 할당된 시간을 저장할 리스트
                print(f"행 {row-1}: 초기 할당 시간: {allocatedHours}")

                # 새로운 로직: 각 일차에 최대 8시간 할당
                remainingHours = totalPlanHours
                print(f"행 {row-1}: 초기 남은 시간: {remainingHours}")

                for day in range(startDay, endDay + 1):
                    print(f"행 {row-1}: 일차: {day}, 남은 시간: {remainingHours}")

                    if remainingHours > 8:
                        allocatedHours[day-1] = 8
                        remainingHours -= 8
                    else:
                        allocatedHours[day-1] = remainingHours
                        break  # 남은 시간이 8시간 이하이면 할당 후 반복문 종료

                    print(f"행 {row-1}: {day}일차 후 할당 시간: {allocatedHours}, 남은 시간: {remainingHours}")

                # 할당된 시간을 테이블에 설정
                for day, hours in enumerate(allocatedHours, start=1):
                    if day >= startDay and day <= endDay:
                        # ID 열이 추가된 이후의 열 인덱스를 조정하여 올바른 위치에 시간을 설정합니다.
                        dayIndex = dailyTrainingTimeIndexes[day - 1] - 1 if id_present else dailyTrainingTimeIndexes[day - 1]
                        print(f"행 {row-1}, {day}일차: 할당 시간: {hours}, 일차 인덱스: {dayIndex}")

                        if dayIndex != -1:  # 유효한 인덱스인 경우에만 설정
                            print(f"행 {row-1}, {day}일차: 인덱스 {dayIndex}에 시간 {hours} 설정")
                            self.tableWidget.setItem(row - 1, dayIndex, QTableWidgetItem(str(int(hours))))

        # 테이블 업데이트 재활성화
        self.tableWidget.setUpdatesEnabled(True)

    # 복호화된 데이터 위젯 표시 메서드
    def displayDecryptedData(self, data):
        if not data or not isinstance(data, list) or not data[0]:
            QMessageBox.warning(self, "데이터 오류", "데이터가 비어 있거나 예상한 형식이 아닙니다.")
            return

        # 첫 번째 항목에서 모든 키를 열 제목으로 사용
        columns = list(data[0].keys())
        self.tableWidget.setColumnCount(len(columns))
        self.tableWidget.setHorizontalHeaderLabels(columns)

        # 데이터 로우 설정
        self.tableWidget.setRowCount(len(data))

        # 데이터 채우기
        for row, item in enumerate(data):
            for col, key in enumerate(columns):
                value = item.get(key, "")
                # JSON의 값이 리스트나 딕셔너리일 수 있으므로, 문자열로 변환
                if isinstance(value, (list, dict)):
                    value = json.dumps(value)
                self.tableWidget.setItem(row, col, QTableWidgetItem(str(value)))

    # 표찰/총기번호 불러오기 메서드
    def loadBadgeAndGunNumber(self):
        fname, _ = QFileDialog.getOpenFileName(self, 'Open file', './', 'hcell files (*.cell)')
        if not fname:
            QMessageBox.warning(self, "파일 선택 오류", "파일이 선택되지 않았습니다.")
            self.badgeAndGunNumberData = None
            return
        self.selectedBadgeAndGunLabel.setText(f"선택한 총기/표찰번호 파일: {fname}")

        try:
            # os.startfile을 사용하여 파일 열기
            os.startfile(fname)
            
            # 테이블 형태로 데이터를 입력할 수 있는 창 생성
            tableDialog = QDialog(self)
            tableDialog.setWindowTitle("데이터 확인 및 입력")
            layout = QVBoxLayout(tableDialog)
            
            # 설명 라벨 추가
            instructionsLabel = QLabel("한셀에서 데이터를 복사하여 아래 텍스트 박스에 붙여넣기 하십시오.")
            layout.addWidget(instructionsLabel)
            
            # 텍스트 에디터 추가
            textEdit = QPlainTextEdit()
            layout.addWidget(textEdit)
            
            # 확인 버튼 추가
            confirmButton = QPushButton("확인", tableDialog)
            confirmButton.clicked.connect(lambda: self.gunprocessCopiedData(textEdit, tableDialog))
            layout.addWidget(confirmButton)
            
            # 다이얼로그 실행
            tableDialog.exec_()

            QMessageBox.information(self, "성공", "파일을 성공적으로 불러왔습니다.")

        except Exception as e:
            QMessageBox.critical(self, "오류 발생", f"파일을 불러오는 데 실패했습니다: {e}")
            print(f"디버깅 정보: {e}")

    def gunprocessCopiedData(self, textEdit, dialog):
        copiedText = textEdit.toPlainText()
        if not copiedText:
            QMessageBox.warning(self, "입력 오류", "데이터가 입력되지 않았습니다.")
            return
        
        # 데이터를 줄 단위로 분리하고 각 줄을 다시 탭 또는 콤마로 분리하여 2차원 리스트로 변환
        self.badgeAndGunNumberData = [line.split('\t') for line in copiedText.split('\n') if line.strip()]
        
        dialog.accept()

    # 전자서명 명부 저장하기 위한 메서드
    def saveFile(self):
        base_dir = 'C:/임시폴더'
        base_filename = '임시파일'
        extension = '.xlsx'
        
        # 폴더가 존재하지 않으면 생성
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)

        # 동일한 파일명이 존재할 경우 파일명을 변경
        counter = 0
        while True:
            if counter == 0:
                temp_filename = os.path.join(base_dir, f'{base_filename}{extension}')
            else:
                temp_filename = os.path.join(base_dir, f'{base_filename}_{counter}{extension}')
            if not os.path.exists(temp_filename):
                break
            counter += 1

        # 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active
        
        # 헤더 저장
        for col in range(self.tableWidget.columnCount()):
            header = self.tableWidget.horizontalHeaderItem(col)
            if header:
                ws.cell(row=1, column=col+1, value=header.text())
        
        # 데이터 및 이미지 저장
        for row in range(self.tableWidget.rowCount()):
            for col in range(self.tableWidget.columnCount()):
                cell = self.tableWidget.item(row, col)
                if cell:
                    value = cell.text()
                    
                    # 이미지 경로 변경
                    if value.endswith('.png') or value.endswith('.gif'):
                        # 경로에서 파일명 추출
                        filename = os.path.basename(value)
                        # 새로운 경로로 파일명 결합
                        new_path = os.path.join(self.signatureSavePath, filename)
                        value = new_path  # 이미지 경로를 무조건 변경된 경로로 설정
                    
                    cell_ref = ws.cell(row=row+2, column=col+1, value=value)
                    
                    # 이미지 파일을 엑셀에 추가 (존재하는 경우에만)
                    if value.endswith('.png') or value.endswith('.gif'):
                        if os.path.exists(value):
                            img = OpenpyxlImage(value)
                            img.anchor = ws.cell(row=row+2, column=col+1).coordinate
                            ws.add_image(img)

                            # 셀 글자색을 흰색으로 변경
                            cell_ref.font = Font(color="FFFFFF")
        
        wb.save(temp_filename)
        wb.close()
        
        # 새 메서드 호출, fname 인자 제거
        self.saveAsHCell(temp_filename)

    # 전자서명 명부 저장하기 한셀 재 저장을 위한 메서드
    def saveAsHCell(self, source_file):
        # 사용자가 파일을 저장할 위치와 이름을 선택할 수 있도록 함
        target_fname, _ = QFileDialog.getSaveFileName(None, 'Save file', './', 'HCell files (*.cell)')
        if target_fname:  # 사용자가 파일 이름을 제공했다면, 처리 계속 진행
            try:
                # 한셀(HCell)로 저장
                hcell = Dispatch('HCell.Application')
                hcell.Visible = True
                workbook = hcell.Workbooks.Open(source_file)

                # HCell 파일 형식으로 저장
                workbook.SaveAs(target_fname, FileFormat=51)
                workbook.Close(False)

                # 저장한 .cell 파일 다시 열기
                workbook = hcell.Workbooks.Open(target_fname)
                worksheet = workbook.Sheets(1)

                try:
                    count = 0
                    # 기존 이미지 정보 저장
                    images_info = {}
                    for shape in worksheet.Shapes:
                        if shape.Type == 13:  # 13은 이미지 타입
                            images_info[shape.TopLeftCell.Address] = shape

                    # 새 이미지 배치
                    total_cells = worksheet.UsedRange.Cells.Count
                    progress_dialog = QProgressDialog("새 이미지 배치 중...", "취소", 0, total_cells)
                    progress_dialog.setWindowTitle("진행 상황")
                    progress_dialog.setWindowModality(Qt.WindowModal)
                    progress_dialog.show()
                    QApplication.processEvents()

                    current_cell = 0
                    for cell in worksheet.UsedRange.Cells:
                        current_cell += 1
                        if cell.Text.endswith(('.png', '.gif')) and os.path.exists(cell.Text):
                            if cell.Address in images_info:
                                existing_shape = images_info[cell.Address]
                                left, top, width, height = existing_shape.Left, existing_shape.Top, existing_shape.Width, existing_shape.Height
                            else:
                                left, top, width, height = cell.Left, cell.Top, cell.Width, cell.Height

                            # 새 이미지 추가
                            shape = worksheet.Shapes.AddPicture(Filename=cell.Text, LinkToFile=False, SaveWithDocument=True, Left=left, Top=top, Width=width, Height=height)
                            shape.Placement = 1  # xlMoveAndSize
                            count += 1
                          
                            print(f"{count}번째 새 이미지가 '{cell.Text}' 경로에서 셀 {cell.Address}에 배치되었습니다. Placement가 1로 설정됨.")

                        # 진행 상황 업데이트
                        progress_dialog.setValue(current_cell)
                        if progress_dialog.wasCanceled():
                            break
                        QApplication.processEvents()

                    # 'Image'로 시작하는 이름을 가진 기존 이미지 객체 삭제
                    for shape in list(worksheet.Shapes):
                        if shape.Name.startswith("Image"):
                            shape.Delete()

                    # 모든 이미지 크기를 한 개의 셀 크기로 조정
                    cell_width = worksheet.Cells(1, 1).Width
                    cell_height = worksheet.Cells(1, 1).Height
                    for shape in worksheet.Shapes:
                        if shape.Type == 13:  # 13은 이미지 타입
                            shape.Width = cell_width
                            shape.Height = cell_height

                    print(f"총 {count}개의 새 이미지가 배치되었으며, 'Image'로 시작하는 모든 오래된 이미지가 삭제되었습니다.")
                except Exception as e:
                    print(f"이미지 로딩 및 배치 중 오류 발생: {e}")

                workbook.Save()
                workbook.Close(False)
                hcell.Quit()

                # 파일 저장 완료 메시지 박스 표시
                QMessageBox.information(None, "저장 완료", "파일이 성공적으로 저장되었습니다!")
            except Exception as e:
                QMessageBox.critical(None, "오류 발생", f"파일 저장 중 오류 발생: {e}")
        else:
            print("파일 저장이 취소되었습니다.")

    # 전자서명 명부 저장하기(이미지 미포함) 관련 메서드
    def saveFileWithoutImages(self):
        base_dir = 'C:/임시폴더'
        base_filename = '임시파일'
        extension = '.xlsx'
        
        # 폴더가 존재하지 않으면 생성
        if not os.path.exists(base_dir):
            os.makedirs(base_dir)

        # 동일한 파일명이 존재할 경우 파일명을 변경
        counter = 0
        while True:
            if counter == 0:
                temp_filename = os.path.join(base_dir, f'{base_filename}{extension}')
            else:
                temp_filename = os.path.join(base_dir, f'{base_filename}_{counter}{extension}')
            if not os.path.exists(temp_filename):
                break
            counter += 1

        # 엑셀 파일 생성
        wb = Workbook()
        ws = wb.active
        
        # 헤더 저장
        for col in range(self.tableWidget.columnCount()):
            header = self.tableWidget.horizontalHeaderItem(col)
            if header:
                ws.cell(row=1, column=col+1, value=header.text())
        
        # 데이터 저장 (이미지 제외)
        for row in range(self.tableWidget.rowCount()):
            for col in range(self.tableWidget.columnCount()):
                cell = self.tableWidget.item(row, col)
                if cell:
                    value = cell.text()
                    
                    # 이미지 경로 변경
                    if value.endswith('.png') or value.endswith('.gif'):
                        # 경로에서 파일명 추출
                        filename = os.path.basename(value)
                        # 새로운 경로로 파일명 결합
                        new_path = os.path.join(self.signatureSavePath, filename)
                        value = new_path  # 이미지 경로를 무조건 변경된 경로로 설정
                    
                    cell_ref = ws.cell(row=row+2, column=col+1, value=value)
                    
                    # 이미지 경로가 있는 경우 셀 글자색을 흰색으로 변경
                    if value.endswith('.png') or value.endswith('.gif'):
                        cell_ref.font = Font(color="FFFFFF")

        
        wb.save(temp_filename)
        wb.close()
        
        # 새 메서드 호출, fname 인자 제거
        self.saveAsHCellWithoutImages(temp_filename)

    # 전자서명 명부 저장하기(이미지 미포함) 한셀 저장 관련 메서드
    def saveAsHCellWithoutImages(self, source_file):
        # 사용자가 파일을 저장할 위치와 이름을 선택할 수 있도록 함
        target_fname, _ = QFileDialog.getSaveFileName(None, 'Save file', './', 'HCell files (*.cell)')
        if target_fname:  # 사용자가 파일 이름을 제공했다면, 처리 계속 진행
            try:
                # 한셀(HCell)로 저장
                hcell = Dispatch('HCell.Application')
                hcell.Visible = True
                workbook = hcell.Workbooks.Open(source_file)
                # HCell 파일 형식으로 저장
                workbook.SaveAs(target_fname, FileFormat=51)
                workbook.Close(False)

                hcell.Quit()

                # 파일이 성공적으로 저장되었음을 알리는 메시지 박스
                QMessageBox.information(self, "성공", "파일이 성공적으로 저장되었습니다!")
            except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {e}")
        else:
            print("파일 저장이 취소되었습니다.")

    # 훈련결산 저장하기 버튼 클릭 이벤트에 연결될 메서드
    def openSaveTrainingSummaryDialog(self):
        day, ok = QInputDialog.getInt(self, "훈련일차 선택", "훈련일차를 선택하세요 (1-5):", minValue=1, maxValue=5)
        if ok:
            dialog = SaveColumnSelectorDialog(self.tableWidget, day, self)
            if dialog.exec_() == QDialog.Accepted:
                selected_columns = dialog.getSelectedColumns()
                # 선택된 열을 처리하는 로직 추가
                print(selected_columns)

    # 훈련결산 종합하기(동미참) 버튼 관련 메서드
    def openTrainingSummaryDialog(self):
        dialog = TrainingSummaryDialog(self)
        dialog.exec_()

    # 중식/교통비 확인(CMS) 버튼 클릭 이벤트에 연결될 메서드 추가
    def openCheckMealTransportDialog(self):
        day, ok = QInputDialog.getInt(self, "훈련일차 선택", "훈련일차를 선택하세요 (1-5):", minValue=1, maxValue=5)
        if ok:
            dialog = CheckMealTransportDialog(self.tableWidget, day, self)
            if dialog.exec_() == QDialog.Accepted:
                selected_columns = dialog.getSelectedColumns()
                # 선택된 열을 처리하는 로직 추가
                print(selected_columns)

    # 전자서명 명부 서식 생성 관련 메서드
    def createSignatureForm(self):
        savePath, _ = QFileDialog.getSaveFileName(self, 'Save file', './', 'HCell files (*.cell)')
        if not savePath:
            return
        
        try:
            hcell = Dispatch('HCell.Application')
            hcell.Visible = True

            workbook = hcell.Workbooks.Add()
            sheet = workbook.Sheets(1)
            sheet.Name = '전자서명 명부'

            headers = [
                '지역대', '훈련시작일차', '훈련종료일차', '훈련명칭', '예비군부대', '계급', '성명', '생년월일', 
                '훈련유형', '개인차수', '총계획시간', '예금주', '은행', '계좌번호', '1일차 훈련시간', '1일차 차감시간', '1일차 차감사유', '1일차 표찰', 
                '1일차 총기번호', '1일차 중식 미신청 여부', '1일차 중식 신청불가', '1일차 교통비 신청불가', '1일차 훈련비 신청불가',
                '1일차 입소서명', '1일차 퇴소서명', '2일차 훈련시간', '2일차 차감시간', '2일차 차감사유', '2일차 표찰', '2일차 총기번호', 
                '2일차 중식 미신청 여부', '2일차 중식 신청불가', '2일차 교통비 신청불가', '2일차 훈련비 신청불가', '2일차 입소서명', 
                '2일차 퇴소서명', '3일차 훈련시간', '3일차 차감시간', '3일차 차감사유', '3일차 표찰', '3일차 총기번호', '3일차 중식 미신청 여부', 
                '3일차 중식 신청불가', '3일차 교통비 신청불가', '3일차 훈련비 신청불가',  '3일차 입소서명', '3일차 퇴소서명', '4일차 훈련시간', '4일차 차감시간', '4일차 차감사유',
                '4일차 표찰', '4일차 총기번호', '4일차 중식 미신청 여부', '4일차 중식 신청불가', '4일차 교통비 신청불가', '4일차 훈련비 신청불가',  
                '4일차 입소서명', '4일차 퇴소서명', '5일차 훈련시간', '5일차 차감시간', '5일차 차감사유', '5일차 표찰', '5일차 총기번호', '5일차 중식 미신청 여부', 
                '5일차 중식 신청불가', '5일차 교통비 신청불가', '5일차 훈련비 신청불가',  '5일차 입소서명', '5일차 퇴소서명'
            ]

            for col_num, header in enumerate(headers, start=1):
                sheet.Cells(1, col_num).Value = header

            workbook.SaveAs(savePath, FileFormat=51)
            workbook.Close(False)
            hcell.Quit()
            QMessageBox.information(self, '성공', '전자서명 명부 서식이 성공적으로 생성되었습니다.')
        except Exception as e:
            QMessageBox.critical(self, '오류 발생', f'서식을 생성하는 동안 오류가 발생했습니다: {e}')

# 여기까지 컨트롤 프레임 관련 메서드
                
# 여기서부터 중간 프레임 관련 메서드
                
 # 검색 관련 여기서부터

    def keyPressEvent(self, event):
        # 엔터 키를 누르면 검색 버튼을 클릭
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            self.searchButton.click()

        # Ctrl+Shift+Z를 누르면 검색 라벨로 포커스 이동
        if event.modifiers() == (Qt.ControlModifier | Qt.ShiftModifier) and event.key() == Qt.Key_Z:
            self.searchLineEdit.setFocus()
            self.searchLineEdit.selectAll()
            
    # 검색 창 관련 메서드
    def searchInTable(self):
        keyword = self.searchLineEdit.text().lower()  # 검색어를 소문자로 변환
        if not keyword:  # 검색어가 비어 있으면 모든 행을 다시 보여주고 검색 인덱스를 초기화
            self.showAllRows()
            self.currentSearchIndex = -1
            return

        self.searchResults = []  # 검색 결과를 저장할 리스트 초기화
        for row in range(self.tableWidget.rowCount()):
            for col in range(self.tableWidget.columnCount()):
                item = self.tableWidget.item(row, col)
                if item and keyword in item.text().lower():  # 검색어와 일치하면 결과 리스트에 추가
                    if row not in self.searchResults:  # 중복 방지
                        self.searchResults.append(row)

        self.showSearchResults()  # 검색 결과 보여주기

    # 검색 결과 보여주는 메서드
    def showSearchResults(self):
        self.hideAllRows()  # 우선 모든 행을 숨깁니다.
        for row in self.searchResults:
            self.tableWidget.showRow(row)  # 검색 결과에 해당하는 행만 표시
        # 검색 결과의 수를 레이블에 업데이트
        self.dataCountLabel.setText(f"데이터 조회 : {len(self.searchResults)}")    

    # 검색 관련 행 숨기기
    def hideAllRows(self):
        for row in range(0, self.tableWidget.rowCount()):  # 첫 번째 행(0 인덱스)을 제외하고 숨김
            self.tableWidget.hideRow(row)

    # 검색 관련 행 보이기
    def showAllRows(self):
        for row in range(self.tableWidget.rowCount()):
            self.tableWidget.showRow(row)
        # 전체 데이터의 수를 레이블에 업데이트
        self.updateRowCount()            
  
  # 검색 관련 여기까지

  # 열 관리 창 관련 여기서부터
   
    # 열 관리 메서드
    def manageColumns(self):
        columnDialog = QDialog(self)  # 열 관리를 위한 다이얼로그 생성
        columnDialog.setWindowTitle('열 관리')
        dialogLayout = QVBoxLayout(columnDialog)  # 다이얼로그의 레이아웃을 설정

        scrollArea = QScrollArea(columnDialog)  # 스크롤 영역 생성
        scrollAreaWidgetContents = QWidget()  # 스크롤 영역에 넣을 내용 위젯
        scrollAreaLayout = QVBoxLayout(scrollAreaWidgetContents)  # 내용 위젯의 레이아웃

        # 헤더의 레이블을 사용하여 체크박스 생성
        for i in range(self.tableWidget.columnCount()):
            # 헤더 아이템의 텍스트를 가져옵니다. 헤더가 없으면 "컬럼 i"를 사용합니다.
            headerItem = self.tableWidget.horizontalHeaderItem(i)
            columnName = headerItem.text() if headerItem else f"컬럼 {i+1}"
            checkBox = QCheckBox(columnName)
            checkBox.setChecked(not self.tableWidget.isColumnHidden(i))
            checkBox.stateChanged.connect(lambda state, x=i: self.toggleColumnVisibility(x, state))
            scrollAreaLayout.addWidget(checkBox)

        scrollAreaWidgetContents.setLayout(scrollAreaLayout)
        scrollArea.setWidget(scrollAreaWidgetContents)
        scrollArea.setWidgetResizable(True)

        dialogLayout.addWidget(scrollArea)
        columnDialog.setLayout(dialogLayout)
        columnDialog.setFixedSize(300, 400)  # 다이얼로그 크기 설정
        columnDialog.exec_()  # 다이얼로그 실행

    # 열 가시성 토글 메서드
    def toggleColumnVisibility(self, column, state):
        self.tableWidget.setColumnHidden(column, not state)            

# 열 관리 창 관련 여기까지

    # 인원추가 버튼 관련 메서드
    def addNewRow(self):
        dialog = AddRowDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            date, name = dialog.getData()
            formattedData = f"{date}_{name}"  # '날짜_성명' 형식으로 포맷
            rowCount = self.tableWidget.rowCount()
            self.tableWidget.insertRow(rowCount)
            self.updateRowCount()  # 새 행 추가 시 카운트 업데이트

            # 새 행의 첫 번째 열에 날짜와 이름 결합 데이터 입력
            self.tableWidget.setItem(rowCount, 0, QTableWidgetItem(formattedData))


    # 데이터 조회 관련 메서드
    def updateRowCount(self):
        # 숨겨지지 않은 행의 수를 계산
        visibleRowCount = sum(1 for _ in range(self.tableWidget.rowCount()) if not self.tableWidget.isRowHidden(_))
        self.dataCountLabel.setText(f"데이터 조회 : {visibleRowCount}")

# 여기까지 중간 프레임 관련 메서드

if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = SignatureApp()
    ex.show()
    sys.exit(app.exec())
